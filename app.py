import os, json, re, tempfile, io
from datetime import datetime
import streamlit as st
import anthropic
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="OM Analyzer", page_icon="🏢", layout="wide")

st.markdown("""
<style>
#MainMenu, footer, header {visibility: hidden;}
html, body, [class*="css"], .stApp, .main {
    background-color: #0D1B2A !important; color: #E0E6EF !important;
}
.block-container {
    padding-top: 0 !important; max-width: 100% !important;
    padding-left: 0 !important; padding-right: 0 !important;
    background: #0D1B2A !important;
}
section[data-testid="stSidebar"] { display: none !important; }
.rv-navbar {
    background: #0B1929; border-bottom: 1px solid #1E3148;
    padding: 0 48px; height: 130px;
    display: flex; align-items: center; justify-content: space-between;
    position: sticky; top: 0; z-index: 999;
}
.rv-logo-block { display: flex; align-items: center; gap: 20px; }
.rv-logo-icon {
    width: 80px; height: 80px; background: #1DC9A4; border-radius: 20px;
    display: flex; align-items: center; justify-content: center;
    font-weight: 800; font-size: 28px; color: #0D1B2A; flex-shrink: 0;
}
.rv-logo-text { display: flex; flex-direction: column; }
.rv-logo-title { font-size: 36px; font-weight: 700; color: #FFFFFF; line-height: 1.2; }
.rv-logo-sub { font-size: 12px; font-weight: 500; color: #5A8FAA; letter-spacing: 0.12em; text-transform: uppercase; margin-top: 6px; }
.rv-nav-right { display: flex; align-items: center; gap: 20px; }
.rv-version { font-size: 12px; color: #3A5A70; }
.rv-claude-badge {
    background: transparent; border: 1.5px solid #1DC9A4; color: #1DC9A4;
    border-radius: 20px; padding: 7px 18px; font-size: 12px; font-weight: 700;
    letter-spacing: 0.05em; display: flex; align-items: center; gap: 6px;
}
.rv-right-panel {
    width: 380px; flex-shrink: 0; background: #091420;
    border-left: 1px solid #1A2E42; padding: 36px 28px;
    margin-right: -5rem;
    min-height: 100vh;
}
.rv-panel-heading {
    font-size: 13px; font-weight: 700; letter-spacing: 0.12em; text-transform: uppercase;
    color: #1DC9A4; margin-bottom: 12px; margin-top: 28px;
}
.rv-panel-heading:first-child { margin-top: 0; }
.rv-bullet { display: flex; align-items: flex-start; gap: 10px; margin-bottom: 10px; }
.rv-bullet-dot { width: 6px; height: 6px; border-radius: 50%; background: #1DC9A4; margin-top: 6px; flex-shrink: 0; }
.rv-bullet-txt { font-size: 14px; color: #5A8FAA; line-height: 1.6; }
.rv-brokers { font-size: 13px; color: #3A6080; line-height: 2.0; margin-top: 6px; }
.rv-meta { font-size: 13px; color: #3A6080; margin-top: 7px; }
.rv-meta span { color: #5A8FAA; }
.rv-divider { border: none; border-top: 1px solid #152030; margin: 20px 0; }

/* Section label with extending line — like Rent Roll */
.rv-section-label-wrap {
    display: flex; align-items: center; gap: 14px; margin-bottom: 18px; margin-top: 0;
}
.rv-section-label {
    font-size: 11px; font-weight: 700; letter-spacing: 0.14em; text-transform: uppercase;
    color: #1DC9A4; white-space: nowrap; flex-shrink: 0;
}
.rv-section-label-line { flex: 1; height: 1px; background: #1A2E42; }

.rv-upload-card {
    background: #0F2438; border: 1px solid #1A3250; border-radius: 14px;
    padding: 32px 34px 26px; margin-bottom: 10px;
}
.rv-upload-title { font-size: 22px; font-weight: 700; color: #FFFFFF; margin-bottom: 8px; }
.rv-upload-sub { font-size: 13px; color: #4A7090; }
.rv-steps { display: flex; gap: 10px; margin-top: 0; }
.rv-step { flex: 1; background: #0F2133; border: 1px solid #1A3250; border-radius: 10px; padding: 18px 20px; }
.rv-step-num { font-size: 10px; font-weight: 700; color: #1DC9A4; text-transform: uppercase; letter-spacing: 0.08em; margin-bottom: 6px; }
.rv-step-txt { font-size: 12px; color: #C0D0E0; line-height: 1.5; }
.rv-file-info {
    background: #0F2133; border: 1px solid #1E3148; border-radius: 8px;
    padding: 10px 16px; font-size: 13px; color: #5A8FAA; margin-bottom: 12px;
}
.rv-file-info b { color: #C0D0E0; }
.rv-success {
    background: #091C11; border: 1px solid #1DC9A4; border-radius: 10px;
    padding: 16px 20px; margin: 16px 0; display: flex; align-items: center; gap: 12px;
}
.rv-success-icon { font-size: 22px; }
.rv-success-text { font-size: 14px; color: #C0D0E0; }
.rv-success-text b { color: #1DC9A4; }

/* Checkboxes in right panel — native Streamlit styled dark */
[data-testid="stCheckbox"] {
    background: transparent !important;
    padding: 2px 4px 2px 0 !important;
    margin-bottom: 2px !important;
}
[data-testid="stCheckbox"] > label {
    display: flex !important;
    align-items: flex-start !important;
    gap: 9px !important;
    cursor: pointer !important;
}
[data-testid="stCheckbox"] p {
    font-size: 13px !important;
    color: #5A8FAA !important;
    line-height: 1.45 !important;
    margin: 0 !important;
}
[data-testid="stCheckbox"]:hover p { color: #8ABDD0 !important; }
/* Checkbox box itself */
[data-testid="stCheckbox"] [data-testid="stWidgetLabel"] { display: none !important; }
[data-testid="stCheckbox"] > label > div:first-child {
    width: 15px !important; height: 15px !important;
    border: 1.5px solid #2A5070 !important;
    border-radius: 3px !important;
    background: #0B1E30 !important;
    flex-shrink: 0 !important;
    margin-top: 2px !important;
}
[data-testid="stCheckbox"] input:checked ~ div {
    background: #1DC9A4 !important;
    border-color: #1DC9A4 !important;
}
[data-testid="stCheckbox"] svg { color: #0D1B2A !important; }
/* Divider between tab groups */
.rv-cb-divider { border-top: 1px solid #152030; margin: 12px 0 10px; }
/* Select/Deselect buttons */
[data-testid="stHorizontalBlock"] [data-testid="stButton"] button {
    background: #0F2133 !important;
    border: 1px solid #1A3250 !important;
    color: #1DC9A4 !important;
    font-size: 11px !important;
    font-weight: 600 !important;
    padding: 5px 0 !important;
    border-radius: 6px !important;
}
[data-testid="column"]:last-child {
    padding-right: 0 !important;
    margin-right: 0 !important;
}
[data-testid="column"]:last-child > div:first-child {
    padding-right: 0 !important;
}
/* Kill Streamlit's outer block container right padding */
.block-container {
    padding-right: 0 !important;
}
section.main > div.block-container {
    padding-right: 0 !important;
}
/* Stretch the last column div to viewport edge */
[data-testid="stHorizontalBlock"] > div:last-child {
    padding-right: 0 !important;
    margin-right: 0 !important;
    flex-shrink: 0 !important;
}
div[data-testid="stFileUploaderDropzoneInput"],
.stFileUploader {
    padding-left: 0 !important;
    padding-right: 0 !important;
}
.stFileUploader > div,
.stFileUploader > div > div,
[data-testid="stFileUploadDropzone"] {
    background: #0B1E30 !important;
    border: 1.5px dashed #2A5070 !important;
    border-radius: 12px !important;
    padding: 22px 32px !important;
    margin: 0 !important;
}
.stFileUploader * { color: #6A9AB8 !important; background: transparent !important; }
.stFileUploader small, .stFileUploader span { color: #3A6080 !important; }
[data-testid="stFileUploadDropzone"] > div { background: transparent !important; border: none !important; }
/* Upload Browse button — white background */
[data-testid="stFileUploaderDropzone"] button,
.stFileUploader button {
    background: #FFFFFF !important;
    color: #0D1B2A !important;
    border: none !important;
    border-radius: 6px !important;
    font-weight: 600 !important;
    font-size: 13px !important;
    padding: 6px 16px !important;
}
[data-testid="stFileUploaderDropzone"] button span,
.stFileUploader button span { color: #0D1B2A !important; }
.stButton > button {
    background: #1DC9A4 !important; color: #0D1B2A !important; border: none !important;
    border-radius: 8px !important; font-weight: 700 !important; font-size: 15px !important;
    padding: 12px 0 !important; width: 100% !important;
}
.stButton > button:hover { background: #18B090 !important; }
.stDownloadButton > button {
    background: #0F2133 !important; color: #1DC9A4 !important; border: 1.5px solid #1DC9A4 !important;
    border-radius: 8px !important; font-weight: 600 !important; font-size: 14px !important;
    padding: 10px 0 !important; width: 100% !important;
}
.stDownloadButton > button:hover { background: #1DC9A420 !important; }
div[data-testid="metric-container"] {
    background: #0F2133 !important; border: 1px solid #1E3148 !important;
    border-radius: 10px !important; padding: 14px 16px !important;
}
div[data-testid="metric-container"] label { color: #5A8FAA !important; font-size: 11px !important; }
div[data-testid="metric-container"] [data-testid="stMetricValue"] { font-size: 20px !important; color: #FFFFFF !important; font-weight: 600 !important; }
.stProgress > div > div { background: #1DC9A4 !important; }
.stAlert, .stSuccess, .stError, .stInfo { background: #0F2133 !important; border-color: #1E3148 !important; color: #C0D0E0 !important; border-radius: 8px !important; }
.stTabs [data-baseweb="tab-list"] { background: transparent !important; border-bottom: 1px solid #1E3148 !important; }
.stTabs [data-baseweb="tab"] { color: #5A8FAA !important; font-size: 13px !important; background: transparent !important; border-radius: 6px 6px 0 0 !important; }
.stTabs [aria-selected="true"] { color: #1DC9A4 !important; border-bottom: 2px solid #1DC9A4 !important; background: #0F2133 !important; }
.stTabs [data-baseweb="tab-panel"] { background: #0D1B2A !important; padding-top: 16px !important; }
.streamlit-expanderHeader { background: #0F2133 !important; color: #C0D0E0 !important; border-radius: 8px !important; border: 1px solid #1E3148 !important; }
.gold-header { background: #1A1A18; color: #D4B07A; padding: 5px 14px; border-radius: 6px; font-size: 12px; font-weight: 600; margin: 16px 0 8px; display: inline-block; }
.flag-warn   { background:#1C1408; border-left:3px solid #D4A054; padding:10px 14px; border-radius:0 6px 6px 0; margin:6px 0; }
.flag-good   { background:#091C11; border-left:3px solid #1DC9A4; padding:10px 14px; border-radius:0 6px 6px 0; margin:6px 0; }
.flag-info   { background:#091525; border-left:3px solid #5A8AC0; padding:10px 14px; border-radius:0 6px 6px 0; margin:6px 0; }
.flag-verify { background:#130E1E; border-left:3px solid #9A7ACA; padding:10px 14px; border-radius:0 6px 6px 0; margin:6px 0; }
.flag-title  { font-size:13px; font-weight:600; margin-bottom:3px; color:#E0E6EF; }
.flag-body   { font-size:12px; color:#7A9AB8; line-height:1.5; }
::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-track { background: #0A1520; }
::-webkit-scrollbar-thumb { background: #1E3148; border-radius: 3px; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — PDF TEXT EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════
def extract_pdf_text(path: str) -> str:
    pages = []
    try:
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text(x_tolerance=3, y_tolerance=3) or ""
                if text:
                    pages.append(f"\n--- PAGE {i+1} ---\n{text}")
                for tbl in (page.extract_tables() or []):
                    if tbl:
                        rows = ["\t".join(str(c or "") for c in row) for row in tbl if row]
                        pages.append("[TABLE]\n" + "\n".join(rows) + "\n[/TABLE]")
        return "\n".join(pages)
    except Exception:
        try:
            from pypdf import PdfReader
            reader = PdfReader(path)
            return "\n".join(
                f"--- PAGE {i+1} ---\n{p.extract_text() or ''}"
                for i, p in enumerate(reader.pages)
            )
        except Exception as e:
            raise RuntimeError(f"Could not read PDF: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — AI ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

SYSTEM = """You are a senior multifamily real estate underwriting analyst with 20+ years of experience.
You read Offering Memoranda from any broker (JLL, CBRE, Marcus & Millichap, Cushman & Wakefield,
Newmark, Colliers, Berkadia, Walker & Dunlop, Eastdil, HFF, and boutique firms) and extract
comprehensive underwriting data.

Return ONLY a single valid JSON object — no markdown fences, no explanation text, nothing else.
For missing fields use null. For empty arrays use [].

SCHEMA:
{
  "broker": {"name": null, "agents": [], "date": null},
  "financing": {
    "offering_type": null,
    "debt_contact": null,
    "notes": null,
    "new_financing": {
      "loan_type": null, "lender": null, "loan_amount": null, "loan_to_value": null,
      "interest_rate": null, "rate_type": null, "amortization_years": null,
      "loan_term_years": null, "interest_only_period": null, "dscr": null,
      "recourse": null, "notes": null
    },
    "assumable_debt": {
      "loan_type": null, "lender": null, "loan_amount": null, "loan_to_value": null,
      "interest_rate": null, "rate_type": null, "amortization_years": null,
      "loan_term_years": null, "interest_only_period": null,
      "origination_date": null, "maturity_date": null,
      "monthly_payment": null, "annual_debt_service": null, "dscr": null,
      "prepayment_penalty": null, "recourse": null, "notes": null
    }
  },
  "property": {
    "name": null, "address": null, "city": null, "state": null, "zip": null,
    "county": null, "msa": null, "units": null, "year_built": null,
    "rentable_sf": null, "avg_unit_sf": null, "buildings": null, "floors": null,
    "acres": null, "density": null, "occupancy_pct": null, "occupancy_date": null,
    "developer": null, "asset_class": null
  },
  "investment": {
    "market_rent": null, "market_rent_psf": null,
    "effective_rent": null, "effective_rent_psf": null,
    "comp_market_rent": null, "comp_effective_rent": null,
    "rent_gap": null, "annual_upside": null,
    "replacement_cost_note": null, "rent_growth_note": null,
    "renovation_tiers": [
      {"name": null, "units": null, "description": null,
       "appliances": null, "cabinets": null, "countertops": null,
       "flooring": null, "backsplash": null, "faucets": null,
       "lighting": null, "sinks": null, "premium": null}
    ],
    "exterior_opportunity": null,
    "additional_income": [
      {"name": null, "category": null, "fee_per_unit_per_month": null,
       "occupancy_assumption": null, "monthly_income": null,
       "current_annual": null, "proforma_annual": null,
       "calculation_detail": null, "notes": null}
    ],
    "amenities": [], "unit_features": [], "highlights": []
  },
  "value_add": {
    "scope": null, "total_cost": null, "cost_per_unit": null, "exterior_capex": null,
    "monthly_premium": null, "annual_premium": null, "roi_pct": null,
    "light_upgrade_items": [],
    "by_floor_plan": [
      {"type": null, "sf": null, "units": null, "inplace_rent": null, "inplace_psf": null,
       "rehab_cost": null, "premium": null, "post_rehab_rent": null, "post_rehab_psf": null}
    ]
  },
  "value_add_levers": [
    {"lever": null, "units": null, "monthly_premium": null, "annual_upside": null, "notes": null}
  ],
  "tax": {
    "parcel_id": null, "assessed_value": null, "millage_city": null,
    "millage_county": null, "millage_total": null, "tax_base": null,
    "solid_waste_fee": null, "total_tax": null, "abatement_program": null,
    "abatement_pct": null, "abatement_term_note": null, "abatement_annual_savings": null,
    "ami_pct": null, "max_allowable_rent": null, "avg_inplace_rent": null,
    "rent_headroom": null, "units_compliant": null, "pct_compliant": null
  },
  "demographics": {
    "pop_1mi": null, "pop_3mi": null, "pop_5mi": null,
    "pop_growth_1mi": null, "pop_growth_3mi": null, "pop_growth_5mi": null,
    "pop_2030_1mi": null, "pop_2030_3mi": null, "pop_2030_5mi": null,
    "median_income_1mi": null, "median_income_3mi": null, "median_income_5mi": null,
    "median_income_2030_1mi": null, "median_income_2030_3mi": null, "median_income_2030_5mi": null,
    "income_growth_1mi": null, "income_growth_3mi": null, "income_growth_5mi": null,
    "renter_pct_1mi": null, "renter_pct_3mi": null, "renter_pct_5mi": null,
    "college_pct_1mi": null, "college_pct_3mi": null, "college_pct_5mi": null,
    "white_collar_pct_1mi": null, "white_collar_pct_3mi": null, "white_collar_pct_5mi": null,
    "home_value": null, "home_value_area": null,
    "crime": null, "school_district": null, "elementary": null,
    "middle": null, "high_school": null,
    "employers": [{"name": null, "drive": null, "employees": null, "sector": null, "notes": null}]
  },
  "unit_mix": [
    {"type": null, "plan": null, "count": null, "pct": null, "sf": null,
     "market_rent": null, "market_psf": null, "eff_rent": null, "eff_psf": null,
     "target_rent": null, "upside": null, "occupied": null, "vacant": null}
  ],
  "utilities": [
    {"name": null, "method": null, "paid_by": null,
     "reimbursement": null, "fee": null, "annual_income": null, "notes": null}
  ],
  "site": {
    "roof": null, "roof_age": null, "exterior": null, "foundation": null,
    "hvac": null, "plumbing": null, "wiring": null, "hot_water": null,
    "washer_dryer": null, "life_safety": null,
    "parking_open": null, "parking_reserved": null, "parking_covered": null,
    "parking_garage": null, "parking_total": null, "parking_ratio": null,
    "reserved_fee": null, "pet_yards": null, "storage": null, "notes": null
  },
  "rent_comps_garden": [
    {"id": null, "name": null, "distance": null, "year_built": null, "rent": null, "notes": null}
  ],
  "rent_comps_townhouse": [
    {"id": null, "name": null, "distance": null, "year_built": null, "rent": null, "notes": null}
  ],
  "rent_comps": [
    {"id": null, "name": null, "address": null, "city_state": null,
     "distance": null, "year_built": null, "units": null, "occupancy": null, "avg_sf": null,
     "comp_type": null, "total_market": null, "total_market_psf": null,
     "total_eff": null, "total_eff_psf": null,
     "by_bed": [
       {"type": null, "units": null, "sf": null,
        "market": null, "market_psf": null, "eff": null, "eff_psf": null}
     ]}
  ],
  "financials": {
    "periods": [],
    "income_lines": [
      {"item": null, "is_total": false, "is_subtotal": false, "is_deduction": false,
       "values": {}, "pct": {}, "note": null}
    ],
    "expense_lines": [
      {"item": null, "is_total": false, "is_subtotal": false,
       "values": {}, "per_unit": {}, "note": null}
    ],
    "noi": {}, "noi_per_unit": {}, "capex": {}, "cffo": {}, "cffo_per_unit": {}, "expense_ratio": {}
  },
  "sale_comps": [
    {"name": null, "address": null, "city_state": null, "date": null,
     "year_built": null, "units": null, "price": null, "ppu": null,
     "ppsf": null, "cap_rate": null, "occupancy": null,
     "buyer": null, "seller": null, "notes": null}
  ],
  "market": {
    "submarket": null, "sub_occupancy": null, "sub_rent": null,
    "sub_growth": null, "metro_inventory": null, "metro_occupancy": null,
    "pipeline": null, "absorption": null, "investment_vol": null,
    "market_summary": null,
    "major_developments": [
      {"name": null, "description": null, "cost": null, "jobs": null, "timeline": null}
    ]
  },
  "affordability": {
    "current_rent": null, "avg_hh_income_3mi": null,
    "monthly_affordability_3x": null, "rent_headroom_3mi": null,
    "avg_hh_income_2030_3mi": null, "monthly_affordability_2030_3x": null,
    "rent_headroom_2030_3mi": null, "income_to_rent_ratio": null, "notes": null
  },
  "insurance": {
    "carrier": null, "annual_premium": null, "per_unit": null,
    "quote_source": null, "notes": null
  },
  "management": {
    "fee_pct": null, "fee_annual": null, "fee_per_unit": null,
    "current_manager": null, "proposed_manager": null, "notes": null
  },
  "replacement_cost": {
    "per_unit": null, "per_sf": null, "total": null, "source": null,
    "land_per_unit": null, "land_total": null,
    "hard_cost_per_sf": null, "hard_cost_per_unit": null, "hard_cost_total": null,
    "soft_cost_pct": null, "soft_cost_per_unit": null, "soft_cost_total": null,
    "direct_replacement_per_unit": null, "direct_replacement_per_sf": null, "direct_replacement_total": null,
    "developer_fee_pct": null, "developer_fee_per_unit": null, "developer_fee_total": null,
    "gc_fee_pct": null, "gc_fee_per_unit": null, "gc_fee_total": null,
    "gross_replacement_per_unit": null, "gross_replacement_per_sf": null, "gross_replacement_total": null,
    "notes": null, "source": null
  },
  "investment_highlights": [
    "string narrative point about this property's investment thesis"
  ],
  "concession_burnoff": {
    "has_concessions": null,
    "monthly_income_current": null,
    "monthly_income_projected": null,
    "burnoff_timeline": null,
    "notes": null
  },
  "flags": [
    {"category": null, "title": null, "detail": null}
  ]
}

CRITICAL EXTRACTION RULES:
1. DEMOGRAPHICS: columns order is 1-mile, 3-mile, 5-mile, County, Metro. Extract all three radii.
   If 1-mile data is not present in the OM, leave 1-mile fields as null.
2. FINANCIAL PERIODS: Extract ALL column headers exactly as they appear (e.g. "T12", "T6 Ann", "T3 Ann", "T1 Ann",
   "Pro Forma YR1", "Year 0", "YR1", "YR2", "FY1", "F-3 Proforma Income", "Current Rents Proforma",
   "2nd Generation Leases Proforma", "Year 2", "Year 3", etc.). Do NOT rename or standardize them.
   If an OM shows multiple proforma scenarios side-by-side (e.g. F-3/Current Rents/2nd Gen), include ALL as periods.
   If an OM shows a 5-year forward analysis table, include Year 2 through Year 5 as additional periods.
   Cap total periods at 7 columns. Every income and expense line must have values for ALL period columns present.
   Mark is_total=true for EGI, Total Expenses, NOI, CFFO rows.
   Mark is_subtotal=true for subtotal rows (Net Rental Income, Rental Collections, Total Controllable, NOI Before Reserves, etc.).
   The note field must contain the full underwriting assumption text from the OM.

3. RENT COMPS: Use rent_comps_garden for garden/flat apartment comps, rent_comps_townhouse for townhouse comps.
   If the OM does not split by type, put all comps in rent_comps (full detail array) only.
   Set comp_type = "Garden", "Townhouse", "Flat", "Mid-Rise", "High-Rise", or as labeled in the OM.
   Include distance and year_built for every comp. Include avg occupancy if shown.
   Include a "Subject" row and "Average" row if the OM shows them.

4. VALUE-ADD: Extract the full floor plan table. If no floor plan table exists (only light upgrade list),
   leave by_floor_plan as [] and populate light_upgrade_items instead.
   For rows with N/A rent/cost, use null — never use 0.
   REVENUE LEVERS: If the OM contains a structured revenue/value-add levers table (e.g. rows listing
   "Continue Interior Renovation", "Push Rents to Market", "Install Package Lockers", "Bulk Cable/Internet",
   "Valet Trash", "Smart Home Tech", "Covered Parking", "Washer/Dryer Equipment" etc.), extract EVERY row
   into value_add_levers with: lever (name), units (unit count), monthly_premium ($/unit/month),
   annual_upside (total annual $ upside), notes. Include the grand total row as the last entry with
   lever="TOTAL ANNUAL REVENUE UPSIDE".

5. TAX: Extract parcel, millage, abatement program, AMI compliance if present.
   If no abatement program exists, set all abatement fields to null.

6. OTHER INCOME: Extract EVERY other income line from the OM into additional_income array.
   Include fees, utility reimbursements, parking, internet, laundry, storage, pet, admin, etc.
   For each: name, category, fee_per_unit_per_month, occupancy_assumption,
   monthly_income, current_annual, proforma_annual, calculation_detail.

7. FLAGS: Generate 6-10 flags relevant to THIS specific property. category = Warning / Opportunity / Verify / Info.
   Base flags on actual data found — occupancy trends, expense anomalies, market catalysts,
   value-add ROI, vacancy assumptions, bad debt, tax risks, financing terms, etc.
   Do NOT fabricate flags for things not mentioned in the OM.

8. FINANCING: Extract offering_type, debt_contact, new_financing, assumable_debt fully.
   If All Cash, set offering_type="All Cash" and leave financing sub-objects null.
   If Free & Clear with a soft quote provided, populate new_financing with the quoted terms.
   Convert percentage strings to decimals (e.g. "72%" → 0.72, "5.75%" → 0.0575).
   For assumable_debt, always extract interest_only_period (e.g. "60 months", "5 years") if stated.

9. SALE COMPS: Extract buyer and seller names if disclosed. Include cap rate, $/unit, $/SF.

10. MARKET: Populate market_summary with a 2-3 sentence narrative.
    Populate major_developments with every named project including cost, jobs, timeline.

11. AFFORDABILITY: Extract rent-to-income table if present. Calculate:
    monthly_affordability_3x = avg_hh_income_3mi / 12 / 3
    rent_headroom = monthly_affordability_3x - current_rent

12. INSURANCE: Carrier, annual premium, per-unit cost, quote source.

13. MANAGEMENT: Fee % of EGI, annual $, per-unit, current and proposed manager names.

14. REPLACEMENT COST: Extract the full table if present. Fields: land_per_unit, land_total, 
    hard_cost_per_sf, hard_cost_per_unit, hard_cost_total, soft_cost_pct, soft_cost_per_unit, soft_cost_total,
    direct_replacement_per_unit, direct_replacement_per_sf, direct_replacement_total,
    developer_fee_pct, developer_fee_per_unit, developer_fee_total,
    gc_fee_pct, gc_fee_per_unit, gc_fee_total,
    gross_replacement_per_unit, gross_replacement_per_sf, gross_replacement_total.
    Also set per_unit = gross_replacement_per_unit and total = gross_replacement_total.
    If no breakdown exists, populate just per_unit, per_sf, total, source, notes.

15. DEMOGRAPHICS: Extract 1-mile, 3-mile and 5-mile data separately. Never mix them.
    If demographic data is not in the OM, set all demographic fields to null.

16. Extract every number that exists anywhere in the OM. Do not skip any table or data page.

17. INVESTMENT HIGHLIGHTS: Extract the 6-10 bullet-point highlights from the executive summary / investment profile 
    into investment_highlights as an array of strings. These are the broker's key selling points.

18. CONCESSION BURNOFF: If the OM contains a concession burn-off analysis or timeline, extract:
    monthly_income_current (current monthly income before burnoff), monthly_income_projected (projected after burnoff),
    burnoff_timeline (e.g. "April-May 2026"), and a notes narrative.

19. COLLECTIONS SUMMARY: If a trailing monthly collections table exists (T-6 or similar), include each month
    as a period column in periods array (e.g. "Jun", "Jul", "Aug", "Sep", "Oct", "Nov") so the operating
    statement correctly reflects the trailing period detail. If trailing months AND proforma columns both exist,
    prioritize the proforma columns but include T-3 or T-2 actuals as well.
"""


def analyze_om(pdf_text: str, api_key: str, progress_cb=None) -> dict:
    import httpx
    MAX = 180_000
    text = pdf_text[:MAX]
    if len(pdf_text) > MAX and progress_cb:
        progress_cb("Large OM — using first 180K characters")
    if progress_cb:
        progress_cb("Sending to Claude AI for analysis...")

    client = anthropic.Anthropic(
        api_key=api_key,
        http_client=httpx.Client(timeout=httpx.Timeout(600.0, connect=30.0))
    )

    def _call(prompt_text, max_tok, system_text):
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=max_tok,
            system=system_text,
            messages=[{"role": "user", "content": prompt_text}]
        )
        raw = resp.content[0].text.strip()
        raw = re.sub(r"^```[a-z]*\n?", "", raw).rstrip("`").strip()
        return raw, resp.stop_reason

    def _parse(raw):
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            pass
        try:
            m = re.search(r"\{.*\}", raw, re.DOTALL)
            if m:
                return json.loads(m.group())
        except json.JSONDecodeError:
            pass
        try:
            return json.loads(_fix_truncated_json(raw))
        except Exception:
            pass
        return None

    raw, stop_reason = _call(
        f"Analyze this OM:\n\n{text}",
        max_tok=32000,
        system_text=SYSTEM
    )
    if progress_cb:
        progress_cb("Parsing extracted data...")

    result = _parse(raw)
    if result is not None:
        return result

    if stop_reason == "max_tokens" or len(raw) > 100:
        if progress_cb:
            progress_cb("Response truncated — retrying completion...")
        raw2, _ = _call(
            f"Complete this truncated JSON so it is fully valid. Output ONLY the completed JSON:\n\n{raw}",
            max_tok=16000,
            system_text="You are a JSON completion assistant. Output ONLY valid JSON, nothing else."
        )
        for candidate in [raw + raw2, raw2]:
            result = _parse(candidate)
            if result is not None:
                return result

    raise ValueError("Could not parse AI response. Try uploading a smaller OM or check your API key.")


def _fix_truncated_json(raw: str) -> str:
    in_string = escape_next = False
    for ch in raw:
        if escape_next: escape_next = False; continue
        if ch == "\\": escape_next = True; continue
        if ch == '"': in_string = not in_string
    if in_string:
        raw += '"'
    opens, in_str, esc = [], False, False
    for ch in raw:
        if esc: esc = False; continue
        if ch == "\\": esc = True; continue
        if ch == '"': in_str = not in_str; continue
        if not in_str:
            if ch in "{[": opens.append("}" if ch == "{" else "]")
            elif ch in "}]" and opens: opens.pop()
    return raw + "".join(reversed(opens))


def _v(val, fmt=None, suffix="", default="N/A"):
    if val is None or val == "": return default
    if fmt == "$":
        try: return f"${float(val):,.0f}{suffix}"
        except: return str(val)
    if fmt == "%":
        try: return f"{float(val):.1f}%"
        except: return str(val)
    if fmt == "n":
        try: return f"{int(float(val)):,}{suffix}"
        except: return str(val)
    return f"{val}{suffix}"

def _pct(val, suffix="", default="N/A"):
    """Format a percentage that may be stored as decimal (0.0464) or whole (4.64).
    Rule: if abs(val) <= 2.0, treat as decimal proportion and multiply by 100."""
    if val is None or val == "": return default
    try:
        f = float(val)
        if abs(f) <= 2.0:
            f = f * 100
        result = f"{f:.2f}%"
        return result + suffix if suffix else result
    except:
        s = str(val).strip()
        return s if s.endswith("%") else f"{s}%"

def _psf(val):
    """Safe PSF formatter — returns '—' for None/zero/non-numeric."""
    if val is None or val == "" or val == "N/A": return "—"
    try:
        f = float(val)
        return "—" if f == 0 else f"${f:.2f}"
    except: return "—"

def _is_num(v):
    """Return True only if v is a real number (not None, '', 'N/A', 'n/a')."""
    if v is None or str(v).strip().lower() in ("", "n/a", "na", "—", "-"): return False
    try: float(str(v).replace("$","").replace(",","")); return True
    except: return False


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — EXCEL REPORT GENERATOR  (3-tab, RealVal color palette)
# ══════════════════════════════════════════════════════════════════════════════

# ── Exact RealVal colors (resolved from template XML) ────────────────────────
C_HDR      = "FF44546A"   # Dark slate blue  — section headers
C_HDR2     = "FF2C3644"   # Darker slate     — cover subtitle / totals
C_HDR3     = "FF295781"   # Dark navy        — NOI / key totals
C_SUB_HDR  = "FFD3D3D3"   # Light gray       — table headers / subsections
C_HDR_TEXT = "FFE7E6E6"   # Near-white       — text on dark backgrounds
C_AMBER    = "FFFFC000"   # Amber            — text on NOI rows
C_BLUE_IN  = "FF0070C0"   # Input blue       — data values
C_LABEL    = "FF44546A"   # Slate            — label text
C_BODY     = "FF3E3E3E"   # Dark gray        — body text
C_WHITE    = "FFFFFFFF"
C_ALT      = "FFEBF0F7"   # Pale blue-gray   — alternating rows
C_SUBTOTAL = "FFDDEBF7"   # Light blue       — subtotal rows
C_WARN     = "FFFFF2CC"   # Yellow           — Warning flags
C_GREEN_L  = "FFD0F0D8"   # Green            — Opportunity flags
C_BLUE_L   = "FFBED2E6"   # Blue             — Info flags
C_PURPLE_L = "FFF4CCCC"   # Pink             — Verify flags
C_BORDER   = "FFB7B7B7"   # Border color


def _fill(c): return PatternFill("solid", fgColor=c)
def _border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _sc(ws, row, col, val, bold=False, bg=None, fg="FF3E3E3E",
        size=9, ha="left", wrap=True, italic=False):
    c = ws.cell(row=row, column=col)
    c.value = val
    c.font = Font(name="Calibri", bold=bold, color=fg, size=size, italic=italic)
    c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.border = _border()
    return c

def _fr(ws, row, ncols, bg):
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=bg)

def _sec(ws, row, title, n=14):
    _fr(ws, row, n, C_HDR)
    _sc(ws, row, 1, title, bold=True, bg=C_HDR, fg=C_HDR_TEXT, size=11)
    ws.row_dimensions[row].height = 22
    return row + 1

def _thdr(ws, row, headers, n=14):
    _fr(ws, row, n, C_SUB_HDR)
    for i, h in enumerate(headers):
        _sc(ws, row, 1 + i, h, bold=True, bg=C_SUB_HDR, fg=C_LABEL, size=9, ha="center")
    ws.row_dimensions[row].height = 18
    return row + 1

def _shdr(ws, row, title, n=14):
    _fr(ws, row, n, C_SUB_HDR)
    _sc(ws, row, 1, title, bold=True, bg=C_SUB_HDR, fg=C_LABEL, size=9)
    ws.row_dimensions[row].height = 17
    return row + 1

def _kv(ws, row, label, value, alt=False, n=14):
    bg = C_ALT if alt else C_WHITE
    _fr(ws, row, n, bg)
    _sc(ws, row, 1, label, bold=True, fg=C_LABEL, bg=bg, size=9)
    _sc(ws, row, 2, str(value) if value else "—", fg=C_BLUE_IN, bg=bg, size=9, wrap=True)
    ws.row_dimensions[row].height = 16
    return row + 1

def _drow(ws, row, vals, alt=False, als=None, h=15, n=None, cs=1):
    bg = C_ALT if alt else C_WHITE
    nc = n or (cs + len(vals) - 1)
    _fr(ws, row, nc, bg)
    for i, v in enumerate(vals):
        ha = als[i] if als and i < len(als) else "left"
        fg = C_BLUE_IN if ha == "right" else C_BODY
        _sc(ws, row, cs + i, str(v) if v is not None else "—", bg=bg, fg=fg, size=9, ha=ha)
    ws.row_dimensions[row].height = h
    return row + 1

def _subtrow(ws, row, vals, n=14):
    _fr(ws, row, n, C_SUBTOTAL)
    for i, v in enumerate(vals):
        ha = "left" if i == 0 else ("left" if i == len(vals) - 1 else "right")
        _sc(ws, row, 1 + i, str(v) if v else "—", bold=True, bg=C_SUBTOTAL, fg=C_LABEL, size=9, ha=ha)
    ws.row_dimensions[row].height = 16
    return row + 1

def _totrow(ws, row, vals, n=14, col=C_HDR2):
    _fr(ws, row, n, col)
    for i, v in enumerate(vals):
        ha = "left" if i == 0 else ("left" if i == len(vals) - 1 else "right")
        _sc(ws, row, 1 + i, str(v) if v else "—", bold=True, bg=col, fg=C_HDR_TEXT, size=9, ha=ha)
    ws.row_dimensions[row].height = 18
    return row + 1

def _noirow(ws, row, vals, n=14):
    _fr(ws, row, n, C_HDR3)
    for i, v in enumerate(vals):
        ha = "left" if i == 0 else ("left" if i == len(vals) - 1 else "right")
        _sc(ws, row, 1 + i, str(v) if v else "—", bold=True, bg=C_HDR3, fg=C_AMBER, size=10, ha=ha)
    ws.row_dimensions[row].height = 20
    return row + 1

def _sp(ws, row): ws.row_dimensions[row].height = 6; return row + 1

def _cover(ws, row, line1, line2, n=14):
    _fr(ws, row, n, C_HDR)
    _sc(ws, row, 1, line1, bold=True, bg=C_HDR, fg=C_HDR_TEXT, size=13)
    ws.row_dimensions[row].height = 28; row += 1
    _fr(ws, row, n, C_HDR2)
    _sc(ws, row, 1, line2, bg=C_HDR2, fg=C_HDR_TEXT, size=9)
    ws.row_dimensions[row].height = 15; row += 1
    return _sp(ws, row)


def build_excel(d: dict, filename: str, sections: dict = None) -> bytes:
    # Default all sections on if not provided
    S = sections or {}
    def _on(key): return S.get(key, True)
    date   = datetime.today().strftime("%B %d, %Y")
    prop   = (d.get("property") or {}).get("name") or "Property"
    broker = (d.get("broker") or {}).get("name") or "N/A"
    pd_    = d.get("property") or {}
    inv    = d.get("investment") or {}
    va     = d.get("value_add") or {}
    tax    = d.get("tax") or {}
    fin    = d.get("financials") or {}
    fin_i  = d.get("financing") or {}
    insur  = d.get("insurance") or {}
    mgmt   = d.get("management") or {}
    repl   = d.get("replacement_cost") or {}
    afford = d.get("affordability") or {}
    demo   = d.get("demographics") or {}
    mkt    = d.get("market") or {}
    umix   = d.get("unit_mix") or []
    utils_ = d.get("utilities") or []
    site   = d.get("site") or {}
    rg     = d.get("rent_comps_garden") or []
    rth    = d.get("rent_comps_townhouse") or []
    rcomps = d.get("rent_comps") or []
    scomps = d.get("sale_comps") or []
    flags  = d.get("flags") or []

    addr = f"{pd_.get('address','')}, {pd_.get('city','')}, {pd_.get('state','')} {pd_.get('zip','')}".strip(", ")
    subtitle = f"{prop}  ·  {addr}  ·  {_v(pd_.get('units'),'n')} Units  ·  {_v(pd_.get('year_built'))} Built  ·  {broker}  ·  {date}"

    wb = Workbook()
    wb.remove(wb.active)

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1 — FINANCIALS
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("Financials")
    ws1.sheet_view.showGridLines = False
    # Column widths scale with number of financial periods (up to 7)
    _n_periods = len((d.get("financials") or {}).get("periods") or [])
    _data_w = 14 if _n_periods >= 6 else (16 if _n_periods >= 4 else 18)
    _notes_w = 30 if _n_periods >= 6 else 36
    _fin_cols = {"A": 32}
    for _i, _c in enumerate("BCDEFGHIJK"):
        if _i < _n_periods:
            _fin_cols[_c] = _data_w
        elif _i == _n_periods:
            _fin_cols[_c] = _notes_w
        else:
            _fin_cols[_c] = 12
    for col, w in _fin_cols.items():
        ws1.column_dimensions[col].width = w

    r = 1
    r = _cover(ws1, r, f"MULTIFAMILY UNDERWRITING REPORT  ·  FINANCIALS", subtitle)

    if _on('deal'):
        # ── A. Deal Summary ───────────────────────────────────────────────────────
        r = _sec(ws1, r, "A.  DEAL SUMMARY & PROPERTY DETAILS")
        agents = ", ".join(
            a.get("name", str(a)) if isinstance(a, dict) else str(a)
            for a in ((d.get("broker") or {}).get("agents") or [])
        ) or "N/A"
        for i, (k, v) in enumerate([
            ("Property Name",      prop),
            ("Address",            addr),
            ("County / MSA",       f"{pd_.get('county') or 'N/A'}  ·  {pd_.get('msa') or 'N/A'}"),
            ("Total Units",        _v(pd_.get("units"), "n")),
            ("Year Built",         _v(pd_.get("year_built"))),
            ("Net Rentable SF",    f"{_v(pd_.get('rentable_sf'), 'n')} SF  (Avg {_v(pd_.get('avg_unit_sf'), 'n')} SF/unit)"),
            ("Buildings / Stories",f"{_v(pd_.get('buildings'), 'n')} buildings  ·  {_v(pd_.get('floors'))} stories"),
            ("Land Area",          f"{_v(pd_.get('acres'))} acres  ({_v(pd_.get('density'))} units/acre)"),
            ("Asset Class",        pd_.get("asset_class")),
            ("Occupancy",          f"{_v(pd_.get('occupancy_pct'), '%')}  as of {pd_.get('occupancy_date') or 'N/A'}"),
            ("Offering Type",      fin_i.get("offering_type")),
            ("Broker / Agents",    f"{broker}  ·  {agents}"),
            ("Debt Contact",       fin_i.get("debt_contact")),
        ]):
            r = _kv(ws1, r, k, v, alt=bool(i % 2))
        r = _sp(ws1, r)

    if _on('unitmix'):
        # ── B. Unit Mix ───────────────────────────────────────────────────────────
        r = _sec(ws1, r, "B.  UNIT MIX")
        if umix:
            r = _thdr(ws1, r, ["Unit Type","Plan","Units","Mix %","SF / Unit",
                                "In-Place Rent","In-Place PSF","Mkt Rent","Mkt PSF","Target Rent","Upside / Unit"])
            ra = ["left","left","center","center","right","right","right","right","right","right","right"]
            for i, u in enumerate(umix):
                r = _drow(ws1, r, [
                    _v(u.get("type","—")), u.get("plan") or "—", _v(u.get("count"),"n"),
                    _pct(u.get("pct")), _v(u.get("sf"),"n"),
                    _v(u.get("market_rent"),"$"), _psf(u.get('market_psf')),
                    _v(u.get("eff_rent"),"$"), _psf(u.get('eff_psf')),
                    _v(u.get("target_rent"),"$"), _v(u.get("upside"),"$"),
                ], alt=bool(i % 2), als=ra)
        else:
            r = _kv(ws1, r, "Note", "No unit mix data found.")
        r = _sp(ws1, r)

        # ── C. Operating Statement ────────────────────────────────────────────────
        periods   = fin.get("periods") or []
    if _on('opstat'):
        inc_lines = fin.get("income_lines") or []
        exp_lines = fin.get("expense_lines") or []

        if periods and (inc_lines or exp_lines):
            p_all = periods[:7]  # support up to 7 columns (e.g. F-3/Current/2ndGen + Year2-5)
            r = _sec(ws1, r, f"C.  OPERATING STATEMENT  ({'  ·  '.join(p_all)})")
            r = _thdr(ws1, r, ["Line Item"] + p_all + ["Underwriting Notes"])
            f6 = ["left"] + ["right"] * len(p_all) + ["left"]

            def _fin_row(item):
                vals = item.get("values") or {}
                return [item.get("item", "—")] + [
                    _v(vals.get(p), "$") if vals.get(p) is not None else "—" for p in p_all
                ] + [item.get("note") or ""]

            # Track which total rows were already rendered inline to avoid duplicates
            rendered_totals = set()

            r = _shdr(ws1, r, "INCOME")
            for i, item in enumerate(inc_lines):
                name = (item.get("item") or "").lower()
                if item.get("is_total"):
                    rendered_totals.add(name)
                    # "NOI before reserves" rows render as subtotal (light blue), not amber NOI
                    if "before reserve" in name or "pre reserve" in name:
                        r = _subtrow(ws1, r, _fin_row(item))
                    else:
                        r = _noirow(ws1, r, _fin_row(item))
                elif item.get("is_subtotal"):
                    r = _subtrow(ws1, r, _fin_row(item))
                else:
                    r = _drow(ws1, r, _fin_row(item), alt=bool(i % 2), als=f6)

            r = _shdr(ws1, r, "EXPENSES")
            for i, item in enumerate(exp_lines):
                name = (item.get("item") or "").lower()
                if item.get("is_total"):
                    rendered_totals.add(name)
                    # "Total expenses pre reserve" / "NOI before reserves" → subtotal
                    if "before reserve" in name or "pre reserve" in name:
                        r = _subtrow(ws1, r, _fin_row(item))
                    else:
                        r = _totrow(ws1, r, _fin_row(item))
                elif item.get("is_subtotal"):
                    r = _subtrow(ws1, r, _fin_row(item))
                else:
                    r = _drow(ws1, r, _fin_row(item), alt=bool(i % 2), als=f6)

            # NOI — only render from fin.noi dict if not already rendered inline
            noi_d = fin.get("noi") or {}
            noi_already = any("net operating income" in t for t in rendered_totals)
            if not noi_already and any(_is_num(noi_d.get(p)) for p in p_all):
                r = _noirow(ws1, r, ["NET OPERATING INCOME"] + [
                    _v(noi_d.get(p), "$") if _is_num(noi_d.get(p)) else "—" for p in p_all
                ] + [""])

            # Capital reserves — only if numeric
            capex_d = fin.get("capex") or {}
            if any(_is_num(capex_d.get(p)) for p in p_all):
                r = _drow(ws1, r, ["  Capital Reserves"] + [
                    _v(capex_d.get(p), "$") if _is_num(capex_d.get(p)) else "—" for p in p_all
                ] + [""], als=f6)

            # CFFO — only render if numeric values exist (guards against Mosby's "N/A" string)
            cffo_d = fin.get("cffo") or {}
            if any(_is_num(cffo_d.get(p)) for p in p_all):
                r = _noirow(ws1, r, ["CASH FLOW FROM OPERATIONS"] + [
                    _v(cffo_d.get(p), "$") if _is_num(cffo_d.get(p)) else "—" for p in p_all
                ] + [""])
        else:
            r = _sec(ws1, r, "C.  OPERATING STATEMENT")
            r = _kv(ws1, r, "Note", "No financial data found in OM.")
        r = _sp(ws1, r)
    if _on('valueadd'):

        # ── D. Value-Add ──────────────────────────────────────────────────────────
        r = _sec(ws1, r, "D.  PROPOSED VALUE-ADD BY FLOOR PLAN")
        plans = va.get("by_floor_plan") or []
        if plans:
            r = _thdr(ws1, r, ["Unit Type","Units","SF","In-Place Rent","In-Place PSF",
                                "Rehab Cost / Unit","Premium / Unit","Post-Rehab Rent","Post-Rehab PSF"])
            ra2 = ["left","center","right","right","right","right","right","right","right"]
            for i, p in enumerate(plans):
                r = _drow(ws1, r, [
                    p.get("type","—"), _v(p.get("units"),"n"), _v(p.get("sf"),"n"),
                    _v(p.get("inplace_rent"),"$"), _psf(p.get('inplace_psf')),
                    _v(p.get("rehab_cost"),"$"), _v(p.get("premium"),"$"),
                    _v(p.get("post_rehab_rent"),"$"), _psf(p.get('post_rehab_psf')),
                ], alt=bool(i % 2), als=ra2)

        # renovation tiers
        tiers = inv.get("renovation_tiers") or []
        if tiers:
            r = _sp(ws1, r)
            r = _shdr(ws1, r, "Renovation Tiers")
            r = _thdr(ws1, r, ["Tier","Units","Appliances","Cabinets","Countertops","Flooring","Premium"])
            for i, t in enumerate(tiers):
                r = _drow(ws1, r, [
                    t.get("name","—"), _v(t.get("units"),"n"),
                    t.get("appliances") or "—", t.get("cabinets") or "—",
                    t.get("countertops") or "—", t.get("flooring") or "—",
                    _v(t.get("premium"),"$") if t.get("premium") else "—",
                ], alt=bool(i % 2))

        # light upgrade items
        light = inv.get("light_upgrade_items") or va.get("light_upgrade_items") or []
        if light:
            r = _sp(ws1, r)
            r = _shdr(ws1, r, "Interior Upgrade Scope")
            for i, item in enumerate(light):
                bg = C_ALT if bool(i % 2) else C_WHITE
                _fr(ws1, r, 14, bg)
                _sc(ws1, r, 1, f"{i+1}.  {item}", bg=bg, fg=C_BODY, size=9)
                ws1.row_dimensions[r].height = 15; r += 1

        r = _sp(ws1, r)
        r = _shdr(ws1, r, "CapEx & ROI Summary")
        for i, (k, v) in enumerate([
            ("Total Renovation Cost",       _v(va.get("total_cost"), "$")),
            ("Cost Per Unit (all-in)",      _v(va.get("cost_per_unit"), "$")),
            ("Exterior / Additional CapEx", _v(va.get("exterior_capex"), "$")),
            ("Monthly Rent Premium",        _v(va.get("monthly_premium"), "$")),
            ("Annual Rent Premium",         _v(va.get("annual_premium"), "$")),
            ("Return on Investment",        f"{va.get('roi_pct') or 'N/A'}%"),
            ("Value-Add Scope",             va.get("scope")),
        ]):
            r = _kv(ws1, r, k, v, alt=bool(i % 2))

        # ── Revenue Upside Levers ─────────────────────────────────────────────────
        levers = d.get("value_add_levers") or []
        if levers:
            r = _sp(ws1, r)
            r = _shdr(ws1, r, "Revenue Upside Levers")
            r = _thdr(ws1, r, ["Value-Add Lever", "Units", "Mo. Premium / Unit", "Annual Upside", "Notes"])
            total_auto = 0
            for i, lv in enumerate(levers):
                is_total = "TOTAL" in (lv.get("lever") or "").upper()
                ann = lv.get("annual_upside")
                try:
                    if not is_total:
                        total_auto += float(ann or 0)
                except Exception:
                    pass
                row_data = [
                    lv.get("lever") or "—",
                    _v(lv.get("units"), "n") if not is_total else "",
                    _v(lv.get("monthly_premium"), "$") if not is_total else "",
                    _v(ann, "$"),
                    lv.get("notes") or "",
                ]
                if is_total:
                    r = _noirow(ws1, r, row_data)
                else:
                    r = _drow(ws1, r, row_data, alt=bool(i % 2),
                              als=["left", "center", "right", "right", "left"])
            if levers and "TOTAL" not in (levers[-1].get("lever") or "").upper() and total_auto > 0:
                r = _noirow(ws1, r, ["TOTAL ANNUAL REVENUE UPSIDE", "", "", _v(total_auto, "$"), ""])

        r = _sp(ws1, r)
    if _on('financing'):

        # ── E. Financing ──────────────────────────────────────────────────────────
        r = _sec(ws1, r, "E.  FINANCING & DEBT TERMS")
        r = _shdr(ws1, r, "Offering & Contact")
        for i, (k, v) in enumerate([
            ("Offering Type", fin_i.get("offering_type")),
            ("Debt Contact",  fin_i.get("debt_contact")),
            ("Notes",         fin_i.get("notes")),
        ]):
            r = _kv(ws1, r, k, v, alt=bool(i % 2))

        nf  = fin_i.get("new_financing") or {}
        asd = fin_i.get("assumable_debt") or {}

        r = _shdr(ws1, r, "New Financing")
        r = _thdr(ws1, r, ["Loan Type","Lender","Loan Amount","LTV","Interest Rate",
                            "Rate Type","Loan Term","Amortization","Interest-Only","DSCR","Recourse","Notes"])
        if any(nf.get(k) for k in ["loan_type","lender","loan_to_value","interest_rate"]):
            r = _drow(ws1, r, [
                nf.get("loan_type") or "—", nf.get("lender") or "—",
                _v(nf.get("loan_amount"), "$"), f"{nf.get('loan_to_value') or 'N/A'}%",
                f"{nf.get('interest_rate') or 'N/A'}%", nf.get("rate_type") or "—",
                nf.get("loan_term_years") or "—", nf.get("amortization_years") or "—",
                nf.get("interest_only_period") or "—", nf.get("dscr") or "—",
                nf.get("recourse") or "—", nf.get("notes") or "—",
            ], als=["left","left","right","right","right","center","center","center","center","center","center","left"])
        else:
            _fr(ws1, r, 14, C_ALT)
            offering = (fin_i.get("offering_type") or "").lower()
            if "all cash" in offering:
                msg = "All Cash offering — no financing available. Buyer must close without debt."
            else:
                msg = "No new financing terms provided in OM. Contact debt broker for quote."
            _sc(ws1, r, 1, msg, bg=C_ALT, fg=C_BODY, size=9, italic=True)
            ws1.row_dimensions[r].height = 15; r += 1

        r = _shdr(ws1, r, "Assumable Debt")
        r = _thdr(ws1, r, ["Loan Type","Lender","Loan Amount","LTV","Interest Rate","Rate Type",
                            "Loan Term","Amortization","Interest-Only","Origination","Maturity","Monthly Pmt","Annual DS","DSCR","Recourse"])
        if any(asd.get(k) for k in ["loan_type","lender","loan_to_value","interest_rate"]):
            r = _drow(ws1, r, [
                asd.get("loan_type") or "—", asd.get("lender") or "—",
                _v(asd.get("loan_amount"), "$"), _pct(asd.get("loan_to_value")),
                _pct(asd.get("interest_rate")), asd.get("rate_type") or "—",
                asd.get("loan_term_years") or "—", asd.get("amortization_years") or "—",
                asd.get("interest_only_period") or "—",
                asd.get("origination_date") or "—", asd.get("maturity_date") or "—",
                _v(asd.get("monthly_payment"), "$"), _v(asd.get("annual_debt_service"), "$"),
                asd.get("dscr") or "—", asd.get("recourse") or "—",
            ], als=["left","left","right","right","right","center","center","center","center","center","center","right","right","center","center"])
        else:
            _fr(ws1, r, 14, C_ALT)
            offering = (fin_i.get("offering_type") or "").lower()
            if "all cash" in offering:
                msg = "All Cash offering — no assumable debt. No existing financing on the property."
            else:
                msg = "None — property offered free and clear. No existing assumable debt."
            _sc(ws1, r, 1, msg, bg=C_ALT, fg=C_BODY, size=9, italic=True)
            ws1.row_dimensions[r].height = 15; r += 1
        r = _sp(ws1, r)
    if _on('flags'):

        # ── F. Underwriting Flags ─────────────────────────────────────────────────
        r = _sec(ws1, r, "F.  UNDERWRITING FLAGS")
        flag_bg = {"Warning": C_WARN, "Opportunity": C_GREEN_L, "Info": C_BLUE_L, "Verify": C_PURPLE_L}
        if flags:
            r = _thdr(ws1, r, ["Category", "Flag Title", "Detail"])
            for f in flags:
                cat = f.get("category", "Info")
                bg  = flag_bg.get(cat, C_WHITE)
                _sc(ws1, r, 1, cat, bold=True, bg=bg, size=9, fg=C_LABEL)
                _sc(ws1, r, 2, f.get("title", ""), bold=True, bg=bg, size=9, fg=C_LABEL)
                _sc(ws1, r, 3, f.get("detail", ""), bg=bg, size=9, fg=C_BODY, wrap=True)
                for col in range(4, 15):
                    ws1.cell(row=r, column=col).fill = PatternFill("solid", fgColor=bg)
                ws1.row_dimensions[r].height = 28; r += 1
        else:
            r = _kv(ws1, r, "Note", "No flags generated.")
        r = _sp(ws1, r)
    if _on('tax'):

        # ── G. Tax ────────────────────────────────────────────────────────────────
        r = _sec(ws1, r, "G.  PROPERTY TAX & TAX ABATEMENT")
        r = _shdr(ws1, r, "Property Tax Detail")
        for i, (k, v) in enumerate([
            ("Parcel ID",             tax.get("parcel_id")),
            ("Assessed Market Value", _v(tax.get("assessed_value"), "$")),
            ("Millage — City",        tax.get("millage_city")),
            ("Millage — County",      tax.get("millage_county")),
            ("Total Millage Rate",    tax.get("millage_total")),
            ("Ad Valorem Tax",        _v(tax.get("tax_base"), "$")),
            ("Solid Waste / Fees",    _v(tax.get("solid_waste_fee"), "$")),
            ("Total Annual Tax Bill", _v(tax.get("total_tax"), "$")),
        ]):
            r = _kv(ws1, r, k, v, alt=bool(i % 2))
        r = _shdr(ws1, r, "Tax Abatement Program")
        for i, (k, v) in enumerate([
            ("Program",               tax.get("abatement_program")),
            ("Abatement %",           f"{tax.get('abatement_pct') or 'N/A'}%"),
            ("Commitment Term",       tax.get("abatement_term_note")),
            ("AMI Requirement",       f"{tax.get('ami_pct') or 'N/A'}% of Area Median Income"),
            ("Annual Tax Savings",    _v(tax.get("abatement_annual_savings"), "$")),
            ("Max Allowable Rent",    _v(tax.get("max_allowable_rent"), "$")),
            ("Avg In-Place Rent",     _v(tax.get("avg_inplace_rent"), "$")),
            ("Headroom / Unit",       _v(tax.get("rent_headroom"), "$")),
        ]):
            r = _kv(ws1, r, k, v, alt=bool(i % 2))
        r = _sp(ws1, r)
    if _on('repl'):

        # ── H. Replacement Cost / Insurance / Management ──────────────────────────
        r = _sec(ws1, r, "H.  REPLACEMENT COST  ·  INSURANCE  ·  MANAGEMENT")
        r = _shdr(ws1, r, "Replacement Cost")
        # If detailed breakdown exists (Grand Preserve style), render full table
        has_detail = any(repl.get(k) for k in ["hard_cost_per_sf","land_per_unit","gross_replacement_per_unit"])
        if has_detail:
            repl_rows = [
                ("Component",                "Per Unit",                                           "Per SF",                                               "Total"),
                ("Land",                     _v(repl.get("land_per_unit"),"$"),                   "—",                                                    _v(repl.get("land_total"),"$")),
                ("Hard Costs",               _v(repl.get("hard_cost_per_unit"),"$"),              _v(repl.get("hard_cost_per_sf"),"$"),                   _v(repl.get("hard_cost_total"),"$")),
                ("Soft Costs",               _v(repl.get("soft_cost_per_unit"),"$"),              f"{repl.get('soft_cost_pct') or '—'}% of hard",         _v(repl.get("soft_cost_total"),"$")),
            ]
            r = _thdr(ws1, r, ["Component","Per Unit","Per SF","Total Cost"])
            for i, row in enumerate(repl_rows[1:]):
                r = _drow(ws1, r, list(row), alt=bool(i%2), als=["left","right","right","right"])
            if repl.get("direct_replacement_per_unit"):
                r = _subtrow(ws1, r, ["Direct Replacement Cost", _v(repl.get("direct_replacement_per_unit"),"$"), _v(repl.get("direct_replacement_per_sf"),"$"), _v(repl.get("direct_replacement_total"),"$")])
            dev_rows = []
            if repl.get("developer_fee_per_unit"):
                dev_rows.append(("Developer Fee", _v(repl.get("developer_fee_per_unit"),"$"), f"{repl.get('developer_fee_pct') or '—'}% of project", _v(repl.get("developer_fee_total"),"$")))
            if repl.get("gc_fee_per_unit"):
                dev_rows.append(("GC Fee",        _v(repl.get("gc_fee_per_unit"),"$"), f"{repl.get('gc_fee_pct') or '—'}% of hard costs", _v(repl.get("gc_fee_total"),"$")))
            for i, row in enumerate(dev_rows):
                r = _drow(ws1, r, list(row), alt=bool(i%2), als=["left","right","right","right"])
            if repl.get("gross_replacement_per_unit"):
                r = _noirow(ws1, r, ["Gross Replacement Cost", _v(repl.get("gross_replacement_per_unit"),"$"), _v(repl.get("gross_replacement_per_sf"),"$"), _v(repl.get("gross_replacement_total"),"$")])
            if repl.get("notes"):
                r = _kv(ws1, r, "Notes", repl.get("notes"))
        else:
            for i, (k, v) in enumerate([
                ("Cost Per Unit",     _v(repl.get("per_unit") or repl.get("gross_replacement_per_unit"), "$")),
                ("Cost Per SF",       _v(repl.get("per_sf") or repl.get("gross_replacement_per_sf"), "$")),
                ("Total Replacement", _v(repl.get("total") or repl.get("gross_replacement_total"), "$")),
                ("Land Per Unit",     _v(repl.get("land_per_unit"), "$")),
                ("Hard Cost Per SF",  _v(repl.get("hard_cost_per_sf"), "$")),
                ("Soft Cost %",       f"{repl.get('soft_cost_pct') or 'N/A'}%"),
                ("Source",            repl.get("source")),
                ("Notes",             repl.get("notes")),
            ]):
                r = _kv(ws1, r, k, v, alt=bool(i % 2))
        r = _shdr(ws1, r, "Insurance")
        for i, (k, v) in enumerate([
            ("Carrier / Provider", insur.get("carrier")),
            ("Annual Premium",     _v(insur.get("annual_premium"), "$")),
            ("Per Unit / Year",    _v(insur.get("per_unit"), "$")),
            ("Quote Source",       insur.get("quote_source")),
            ("Notes",              insur.get("notes")),
        ]):
            r = _kv(ws1, r, k, v, alt=bool(i % 2))
        r = _shdr(ws1, r, "Property Management")
        for i, (k, v) in enumerate([
            ("Management Fee %",   f"{mgmt.get('fee_pct') or 'N/A'}% of EGI"),
            ("Annual Fee",         _v(mgmt.get("fee_annual"), "$")),
            ("Per Unit / Year",    _v(mgmt.get("fee_per_unit"), "$")),
            ("Current Manager",    mgmt.get("current_manager")),
            ("Proposed Manager",   mgmt.get("proposed_manager")),
            ("Notes",              mgmt.get("notes")),
        ]):
            r = _kv(ws1, r, k, v, alt=bool(i % 2))
        r = _sp(ws1, r)

        # Disclaimer
        _fr(ws1, r, 14, C_ALT)
        _sc(ws1, r, 1, "AI-generated from broker OM. Internal use only. Verify all figures independently. Powered by Anthropic Claude.",
            bg=C_ALT, fg="FF888880", size=8, italic=True)
        ws1.row_dimensions[r].height = 13

        # ══════════════════════════════════════════════════════════════════════════
    # TAB 2 — COMPARABLES
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Comparables")
    ws2.sheet_view.showGridLines = False
    for col, w in {"A":30,"B":14,"C":10,"D":10,"E":10,"F":12,
                   "G":12,"H":12,"I":12,"J":10,"K":32}.items():
        ws2.column_dimensions[col].width = w

    r = 1
    r = _cover(ws2, r, f"COMPARABLES — RENT & SALE  |  {prop}", subtitle, n=11)

    # ── Comp subject baseline: avg eff_rent across all unit mix rows ──────────
    ra_c = ["left","center","center","center","center","center","right","right","right","right","left"]
    def _subj_rent_for(comp_type):
        """Return avg effective rent for matching unit type, or overall avg."""
        ct = (comp_type or "").lower()
        matching = [u for u in umix if ct in (u.get("plan") or u.get("type") or "").lower()]
        pool = matching if matching else umix
        rents = [float(u["eff_rent"]) for u in pool if u.get("eff_rent") and _is_num(u.get("eff_rent"))]
        return int(sum(rents) / len(rents)) if rents else 0

    def _render_comp_group(ws, r, comps, section_title, comp_type_label, n=11):
        subj_rent = _subj_rent_for(comp_type_label)
        r = _sec(ws, r, section_title, n=n)
        r = _thdr(ws, r, ["Property","Type","Yr Built","Distance","Units",
                           "Occ %","Market Rent","Eff Rent","Rent PSF","vs. Subject","Notes"], n=n)
        for i, c in enumerate(comps):
            rent_val = c.get("rent") or c.get("total_market") or c.get("total_eff")
            try: rent_num = float(str(rent_val).replace("$","").replace(",",""))
            except: rent_num = 0
            vs = f"{int(rent_num - subj_rent):+,}" if rent_num and subj_rent else "—"
            name = c.get("name","—")
            is_highlight = any(k in name.lower() for k in ("subject","average","avg","index"))
            row_data = [name, c.get("comp_type") or comp_type_label,
                        _v(c.get("year_built")), c.get("distance") or "—",
                        _v(c.get("units"),"n") if c.get("units") else "—",
                        _pct(c.get("occupancy")) if c.get("occupancy") else "—",
                        _v(rent_val,"$"), _v(c.get("total_eff") or rent_val,"$"),
                        _psf(c.get("total_market_psf") or c.get("total_eff_psf")),
                        vs, c.get("notes") or "—"]
            if is_highlight:
                r = _totrow(ws, r, row_data, n=n)
            else:
                r = _drow(ws, r, row_data, alt=bool(i % 2), als=ra_c, n=n)
        r = _sp(ws, r)
        return r

    if _on('rentcomps'):
        # ── A. Garden Rent Comps ──────────────────────────────────────────────────
        if rg:
            r = _render_comp_group(ws2, r, rg, "A.  GARDEN RENT COMPARABLES", "Garden")

    if _on('rentcomps'):
        # ── B. Townhouse Rent Comps ───────────────────────────────────────────────
        if rth:
            sec_letter = "B" if rg else "A"
            r = _render_comp_group(ws2, r, rth, f"{sec_letter}.  TOWNHOUSE RENT COMPARABLES", "Townhouse")

    if _on('rentcomps'):
        # ── C. Full Comp Detail (covers Mosby-style OMs with only rcomps) ─────────
        if rcomps:
            sec_full = "A" if (not rg and not rth) else ("C" if (rg and rth) else "B")
            r = _sec(ws2, r, f"{sec_full}.  FULL COMPARABLE DETAIL", n=11)
            r = _thdr(ws2, r, ["#","Property","Type","Yr Built","Distance","Units",
                                "Occ %","Mkt Rent","Eff Rent","Avg SF","Notes"], n=11)
            for i, rc in enumerate(rcomps):
                r = _drow(ws2, r, [
                    rc.get("id","—"), rc.get("name","—"), rc.get("comp_type") or "—",
                    _v(rc.get("year_built")), rc.get("distance") or "—",
                    _v(rc.get("units"),"n"), _pct(rc.get("occupancy")),
                    _v(rc.get("total_market"),"$"), _v(rc.get("total_eff"),"$"),
                    _v(rc.get("avg_sf"),"n"), "—",
                ], alt=bool(i % 2), als=["center","left","center","center","center","center",
                                          "center","right","right","right","left"], n=11)
            r = _sp(ws2, r)

        if not rg and not rth and not rcomps:
            r = _sec(ws2, r, "A.  RENT COMPARABLES", n=11)
            r = _kv(ws2, r, "Note", "No rent comparable data found in OM. Source from listing broker or CoStar.", n=11)
            r = _sp(ws2, r)

    if _on('salecomps'):
        # ── D. Sale Comps ─────────────────────────────────────────────────────────
        r = _sec(ws2, r, "D.  SALE COMPARABLES", n=11)
        if scomps:
            r = _thdr(ws2, r, ["Property","Date","Yr Built","Units","Sale Price",
                                "$/Unit","$/SF","Cap Rate","Occ","Buyer","Seller"], n=11)
            for i, sc in enumerate(scomps):
                r = _drow(ws2, r, [
                    sc.get("name","—"), sc.get("date") or "—", _v(sc.get("year_built")),
                    _v(sc.get("units"),"n"), _v(sc.get("price"),"$"),
                    _v(sc.get("ppu"),"$"), _v(sc.get("ppsf"),"$"),
                    sc.get("cap_rate") or "—", sc.get("occupancy") or "—",
                    sc.get("buyer") or "—", sc.get("seller") or "—",
                ], alt=bool(i % 2), n=11,
                   als=["left","center","center","center","right","right","right","center","center","left","left"])
        else:
            _fr(ws2, r, 11, C_ALT)
            _sc(ws2, r, 1, "Sale comps not provided in OM. Source from CoStar, RCA, or listing broker.",
                bg=C_ALT, fg=C_BODY, size=9, italic=True, wrap=True)
            ws2.row_dimensions[r].height = 20; r += 1
        r = _sp(ws2, r)

    # Disclaimer
    _fr(ws2, r, 11, C_ALT)
    _sc(ws2, r, 1, "AI-generated. Internal use only. Verify all figures independently. Powered by Anthropic Claude.",
        bg=C_ALT, fg="FF888880", size=8, italic=True)
    ws2.row_dimensions[r].height = 13

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 3 — DEMOGRAPHICS
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Demographics")
    ws3.sheet_view.showGridLines = False
    for col, w in {"A":36,"B":20,"C":20,"D":20,"E":14,"F":42}.items():
        ws3.column_dimensions[col].width = w

    r = 1
    r = _cover(ws3, r, f"DEMOGRAPHICS & MARKET OVERVIEW  |  {prop}", subtitle, n=6)

    if _on('addinc'):
        # ── A. Additional Income ──────────────────────────────────────────────────
        add_inc = inv.get("additional_income") or []
        if add_inc:
            r = _sec(ws3, r, "A.  ADDITIONAL INCOME", n=6)
            r = _thdr(ws3, r, ["Income Item","Category","Fee/Unit/Mo","Occ %",
                                "Monthly Income","Current Annual","Pro Forma Annual","Calculation Detail"], n=6)
            for i, inc in enumerate(add_inc):
                r = _drow(ws3, r, [
                    inc.get("name") or "—", inc.get("category") or "—",
                    _v(inc.get("fee_per_unit_per_month"), "$") if inc.get("fee_per_unit_per_month") else "—",
                    inc.get("occupancy_assumption") or "—",
                    _v(inc.get("monthly_income"), "$") if inc.get("monthly_income") else "—",
                    _v(inc.get("current_annual"), "$") if inc.get("current_annual") else "—",
                    _v(inc.get("proforma_annual"), "$") if inc.get("proforma_annual") else "—",
                    inc.get("calculation_detail") or inc.get("notes") or "—",
                ], alt=bool(i % 2), n=6,
                   als=["left","left","right","center","right","right","right","left"])
            r = _sp(ws3, r)

    if _on('utilities'):
        # ── B. Utilities ──────────────────────────────────────────────────────────
        r = _sec(ws3, r, "B.  UTILITY INFORMATION", n=6)
        if utils_:
            r = _thdr(ws3, r, ["Utility","Billing Method","Paid By","Reimbursement","Annual Income","Notes"], n=6)
            for i, u in enumerate(utils_):
                r = _drow(ws3, r, [
                    u.get("name","—"), u.get("method") or "—", u.get("paid_by") or "—",
                    u.get("reimbursement") or "N/A",
                    _v(u.get("annual_income"), "$"), u.get("notes") or "—",
                ], alt=bool(i % 2), als=["left","center","center","left","right","left"], n=6)
        else:
            r = _kv(ws3, r, "Note", "No utility data found in OM.", n=6)
        r = _sp(ws3, r)

    if _on('pop'):
        # ── C. Population & Income ────────────────────────────────────────────────
        r = _sec(ws3, r, "C.  POPULATION & INCOME DEMOGRAPHICS", n=6)
        has_demo = any(demo.get(k) for k in ["pop_1mi","pop_3mi","pop_5mi","median_income_3mi"])
        if not has_demo:
            _fr(ws3, r, 6, C_WARN)
            _sc(ws3, r, 1, "Detailed 1-mile / 3-mile / 5-mile radius demographic data not provided in this OM. Source independently from CoStar, Esri, or census.gov.",
                bg=C_WARN, fg=C_LABEL, size=9, bold=True, wrap=True)
            ws3.row_dimensions[r].height = 30; r += 1; r = _sp(ws3, r)

        r = _thdr(ws3, r, ["Metric","1-Mile Radius","3-Mile Radius","5-Mile Radius","Notes"], n=6)
        for i, (metric, v1, v3, v5, note) in enumerate([
            ("Population (2025)",        _v(demo.get("pop_1mi"),"n"),  _v(demo.get("pop_3mi"),"n"),  _v(demo.get("pop_5mi"),"n"),  ""),
            ("Population (2030 Proj.)",  _v(demo.get("pop_2030_1mi"),"n"), _v(demo.get("pop_2030_3mi"),"n"), _v(demo.get("pop_2030_5mi"),"n"), ""),
            ("Population Growth (5-yr)", demo.get("pop_growth_1mi") or "—", demo.get("pop_growth_3mi") or "—", demo.get("pop_growth_5mi") or "—", ""),
            ("Median HH Income (2025)",  _v(demo.get("median_income_1mi"),"$"), _v(demo.get("median_income_3mi"),"$"), _v(demo.get("median_income_5mi"),"$"), ""),
            ("Median HH Income (2030)",  _v(demo.get("median_income_2030_1mi"),"$"), _v(demo.get("median_income_2030_3mi"),"$"), _v(demo.get("median_income_2030_5mi"),"$"), ""),
            ("Income Growth (5-yr)",     demo.get("income_growth_1mi") or "—", demo.get("income_growth_3mi") or "—", demo.get("income_growth_5mi") or "—", ""),
            ("Renter-Occupied Units",    demo.get("renter_pct_1mi") or "—", demo.get("renter_pct_3mi") or "—", demo.get("renter_pct_5mi") or "—", ""),
            ("Bachelor's Degree+",       demo.get("college_pct_1mi") or "—", demo.get("college_pct_3mi") or "—", demo.get("college_pct_5mi") or "—", ""),
            ("White-Collar Workers",     demo.get("white_collar_pct_1mi") or "—", demo.get("white_collar_pct_3mi") or "—", demo.get("white_collar_pct_5mi") or "—", ""),
            ("Avg Home Value",           "—", _v(demo.get("home_value"),"$"), "—", ""),
        ]):
            bg = C_ALT if bool(i % 2) else C_WHITE
            _fr(ws3, r, 6, bg)
            _sc(ws3, r, 1, metric, bold=True, fg=C_LABEL, bg=bg, size=9)
            _sc(ws3, r, 2, v1, fg=C_BLUE_IN, bg=bg, size=9, ha="right")
            _sc(ws3, r, 3, v3, fg=C_BLUE_IN, bg=bg, size=9, ha="right")
            _sc(ws3, r, 4, v5, fg=C_BLUE_IN, bg=bg, size=9, ha="right")
            _sc(ws3, r, 5, note, fg=C_BODY, bg=bg, size=9, wrap=True)
            ws3.cell(row=r, column=6).fill = PatternFill("solid", fgColor=bg)
            ws3.row_dimensions[r].height = 16; r += 1
        r = _sp(ws3, r)

    if _on('afford'):
        # ── D. Affordability ──────────────────────────────────────────────────────
        r = _sec(ws3, r, "D.  AFFORDABILITY & RENT GROWTH RUNWAY", n=6)
        r = _thdr(ws3, r, ["Metric","2025 (3-Mile)","2030 Proj. (3-Mile)","Notes"], n=6)
        for i, (metric, v25, v30, note) in enumerate([
            ("Current In-Place Rent",          _v(afford.get("current_rent"),"$"), "—", "Subject effective rent"),
            ("Avg HH Income",                  _v(afford.get("avg_hh_income_3mi"),"$"), _v(afford.get("avg_hh_income_2030_3mi"),"$"), ""),
            ("Monthly Affordability (3× rule)",_v(afford.get("monthly_affordability_3x"),"$"), _v(afford.get("monthly_affordability_2030_3x"),"$"), "Income ÷ 12 ÷ 3"),
            ("Rent Headroom",                  _v(afford.get("rent_headroom_3mi"),"$"), _v(afford.get("rent_headroom_2030_3mi"),"$"), "Threshold minus current rent"),
            ("Income-to-Rent Ratio",           afford.get("income_to_rent_ratio") or "—", "—", ""),
        ]):
            bg = C_ALT if bool(i % 2) else C_WHITE
            _fr(ws3, r, 6, bg)
            _sc(ws3, r, 1, metric, bold=True, fg=C_LABEL, bg=bg, size=9)
            _sc(ws3, r, 2, v25, fg=C_BLUE_IN, bg=bg, size=9, ha="right")
            _sc(ws3, r, 3, v30, fg=C_BLUE_IN, bg=bg, size=9, ha="right")
            _sc(ws3, r, 4, note, fg=C_BODY, bg=bg, size=9)
            ws3.cell(row=r, column=5).fill = PatternFill("solid", fgColor=bg)
            ws3.cell(row=r, column=6).fill = PatternFill("solid", fgColor=bg)
            ws3.row_dimensions[r].height = 16; r += 1
        r = _sp(ws3, r)

    if _on('schools'):
        # ── E. Schools & Crime ────────────────────────────────────────────────────
        r = _sec(ws3, r, "E.  SCHOOLS, CRIME & QUALITY OF LIFE", n=6)
        r = _shdr(ws3, r, "Assigned Schools  (source: greatschools.org)", n=6)
        for i, (k, v) in enumerate([
            ("School District", demo.get("school_district") or "Not provided — verify via district map"),
            ("Elementary",      demo.get("elementary") or "Not provided — verify via district map"),
            ("Middle School",   demo.get("middle") or "Not provided — verify via district map"),
            ("High School",     demo.get("high_school") or "Not provided — verify via district map"),
        ]):
            r = _kv(ws3, r, k, v, alt=bool(i % 2), n=6)
        r = _shdr(ws3, r, "Crime  (source: crimegrade.org)", n=6)
        r = _kv(ws3, r, "Crime Data", demo.get("crime") or "Not provided in OM — source independently at crimegrade.org", n=6)
        r = _sp(ws3, r)

    if _on('utilities'):
        # ── F. Site Info ──────────────────────────────────────────────────────────
        r = _sec(ws3, r, "F.  SITE & CONSTRUCTION INFORMATION", n=6)
        r = _shdr(ws3, r, "Physical Plant", n=6)
        for i, (k, v) in enumerate([
            ("Roof / Age",     f"{site.get('roof') or 'N/A'}  —  {site.get('roof_age') or 'N/A'}"),
            ("Exterior",       site.get("exterior")),
            ("Foundation",     site.get("foundation")),
            ("HVAC",           site.get("hvac")),
            ("Plumbing",       site.get("plumbing")),
            ("Wiring",         site.get("wiring")),
            ("Hot Water",      site.get("hot_water")),
            ("Washer / Dryer", site.get("washer_dryer")),
            ("Life Safety",    site.get("life_safety") or "Verify — not specified"),
            ("Construction",   site.get("notes")),
        ]):
            if v: r = _kv(ws3, r, k, v, alt=bool(i % 2), n=6)
        r = _shdr(ws3, r, "Parking & Site Features", n=6)
        for i, (k, v) in enumerate([
            ("Open Spaces",  _v(site.get("parking_open"), "n")),
            ("Covered",      _v(site.get("parking_covered"), "n")),
            ("Garage",       site.get("parking_garage") or "None"),
            ("Total / Ratio",f"{_v(site.get('parking_total'), 'n')} ({site.get('parking_ratio') or 'N/A'})"),
            ("Pet Yards",    site.get("pet_yards") or "N/A"),
            ("Storage",      site.get("storage") or "N/A"),
        ]):
            r = _kv(ws3, r, k, v, alt=bool(i % 2), n=6)
        r = _sp(ws3, r)

    if _on('employers'):
        # ── G. Major Employers ────────────────────────────────────────────────────
        employers = demo.get("employers") or []
        r = _sec(ws3, r, "G.  MAJOR EMPLOYERS & ECONOMIC DRIVERS", n=6)
        if employers and any(e.get("name") for e in employers):
            r = _thdr(ws3, r, ["Employer / Institution","Drive Time","Employees","Sector","Notes"], n=6)
            for i, e in enumerate(employers):
                if e.get("name"):
                    r = _drow(ws3, r, [
                        e.get("name","—"), e.get("drive") or "—", e.get("employees") or "—",
                        e.get("sector") or "—", e.get("notes") or "—",
                    ], alt=bool(i % 2), als=["left","center","center","left","left"], h=22, n=6)
        else:
            r = _kv(ws3, r, "Note", "No employer data found in OM — source from CoStar or broker.", n=6)
        r = _sp(ws3, r)

    if _on('market'):
        # ── H. Market & Submarket Overview ────────────────────────────────────────
        r = _sec(ws3, r, "H.  MARKET & SUBMARKET OVERVIEW", n=6)
        if mkt.get("market_summary"):
            r = _shdr(ws3, r, "Market Narrative", n=6)
            _fr(ws3, r, 6, C_WHITE)
            _sc(ws3, r, 1, mkt["market_summary"], bg=C_WHITE, fg=C_BODY, size=9, wrap=True)
            ws3.row_dimensions[r].height = 45; r += 1; r = _sp(ws3, r)

        r = _shdr(ws3, r, "Submarket Metrics", n=6)
        for i, (k, v) in enumerate([
            ("Submarket",         mkt.get("submarket")),
            ("Sub. Occupancy",    mkt.get("sub_occupancy")),
            ("Sub. Avg Rent",     _v(mkt.get("sub_rent"), "$")),
            ("Sub. Rent Growth",  mkt.get("sub_growth")),
            ("Metro Inventory",   mkt.get("metro_inventory")),
            ("Metro Occupancy",   mkt.get("metro_occupancy")),
        ]):
            if v: r = _kv(ws3, r, k, v, alt=bool(i % 2), n=6)

    if _on('market'):
        # ── I. Supply & Demand ────────────────────────────────────────────────────
        r = _sp(ws3, r)
        r = _sec(ws3, r, "I.  SUPPLY & DEMAND", n=6)
        r = _thdr(ws3, r, ["Metric","Value","Notes"], n=6)
        supply_rows = [
            ("Pipeline Units (Under Construction)", _v(mkt.get("pipeline"), "n"),    "Units currently under construction in submarket"),
            ("Absorption (Units / Year)",           _v(mkt.get("absorption"), "n"),   "Annual net absorption in submarket"),
            ("Investment Volume",                   mkt.get("investment_vol") or "—", "Total multifamily investment volume"),
        ]
        for i, (k, v, note) in enumerate(supply_rows):
            if v and v != "N/A":
                bg = C_ALT if bool(i % 2) else C_WHITE
                _fr(ws3, r, 6, bg)
                _sc(ws3, r, 1, k,    bold=True, fg=C_LABEL,   bg=bg, size=9)
                _sc(ws3, r, 2, v,    fg=C_BLUE_IN, bg=bg, size=9, ha="right")
                _sc(ws3, r, 3, note, fg=C_BODY,    bg=bg, size=9, wrap=True)
                ws3.cell(row=r, column=4).fill = PatternFill("solid", fgColor=bg)
                ws3.cell(row=r, column=5).fill = PatternFill("solid", fgColor=bg)
                ws3.cell(row=r, column=6).fill = PatternFill("solid", fgColor=bg)
                ws3.row_dimensions[r].height = 16; r += 1

        devs = mkt.get("major_developments") or []
        if devs and any(d.get("name") for d in devs):
            r = _sp(ws3, r)
            r = _shdr(ws3, r, "Pipeline Developments", n=6)
            r = _thdr(ws3, r, ["Development","Description","Est. Cost","Jobs","Timeline"], n=6)
            for i, dev in enumerate(devs):
                if dev.get("name"):
                    r = _drow(ws3, r, [
                        dev.get("name","—"), dev.get("description") or "—",
                        dev.get("cost") or "—", dev.get("jobs") or "—",
                        dev.get("timeline") or "—",
                    ], alt=bool(i % 2), als=["left","left","right","center","left"], h=25, n=6)
        r = _sp(ws3, r)

    # Disclaimer
    _fr(ws3, r, 6, C_ALT)
    _sc(ws3, r, 1, "AI-generated. Internal use only. Verify all figures independently. Powered by Anthropic Claude.",
        bg=C_ALT, fg="FF888880", size=8, italic=True)
    ws3.row_dimensions[r].height = 13

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    pass  # hidden via CSS

st.markdown("""
<div class="rv-navbar">
  <div class="rv-logo-block">
    <div class="rv-logo-icon">OM</div>
    <div class="rv-logo-text">
      <div class="rv-logo-title">OM Analyzer</div>
      <div class="rv-logo-sub">RealVal &nbsp;·&nbsp; Multifamily Underwriting Intelligence</div>
    </div>
  </div>
  <div class="rv-nav-right">
    <span class="rv-version">v3.0</span>
    <div class="rv-claude-badge">⚡ CLAUDE AI</div>
  </div>
</div>
""", unsafe_allow_html=True)

main_col, right_col = st.columns([2.2, 1], gap="small")

with right_col:
    # Style the entire right column as the panel
    st.markdown("""
<style>
/* Right column = the panel */
[data-testid="stHorizontalBlock"] > div:last-child {
    background: #091420 !important;
    border-left: 1px solid #1A2E42 !important;
    padding: 28px 20px 28px 20px !important;
    margin-right: -5rem !important;
    min-height: 100vh !important;
}
/* Checkboxes */
[data-testid="stCheckbox"] {
    background: transparent !important;
    padding: 1px 0 !important;
    margin-bottom: 1px !important;
}
[data-testid="stCheckbox"] p {
    font-size: 13px !important;
    color: #5A8FAA !important;
    line-height: 1.4 !important;
}
[data-testid="stCheckbox"]:hover p { color: #8ABDD0 !important; }
[data-testid="stCheckbox"] input:checked + div,
[data-testid="stCheckbox"] input:checked ~ div { background: #1DC9A4 !important; border-color: #1DC9A4 !important; }
[data-testid="stCheckbox"] svg { color: #0D1B2A !important; }
/* Select/Deselect buttons */
[data-testid="column"] [data-testid="stButton"] > button {
    background: #0F2133 !important; border: 1px solid #1A3250 !important;
    color: #1DC9A4 !important; font-size: 11px !important;
    font-weight: 600 !important; padding: 5px 0 !important;
    border-radius: 6px !important; margin-top: 4px !important;
}
[data-testid="column"] [data-testid="stButton"] > button:hover { background: #1DC9A420 !important; }
/* Divider between sections */
.rv-cb-divider { border-top: 1px solid #152030; margin: 10px 0 8px; }
</style>
""", unsafe_allow_html=True)

    st.markdown('<div class="rv-panel-heading" style="margin-top:0;">What\'s in the Report</div>', unsafe_allow_html=True)
    st.markdown('<div class="rv-panel-heading" style="margin-top:10px;">Tab 1 — Financials</div>', unsafe_allow_html=True)

    sel_deal      = st.checkbox("Deal summary & property details",          value=True, key="sel_deal")
    sel_unitmix   = st.checkbox("Unit mix with rent upside",                value=True, key="sel_unitmix")
    sel_opstat    = st.checkbox("Operating statement (all periods)",        value=True, key="sel_opstat")
    sel_valueadd  = st.checkbox("Value-add by floor plan & revenue levers", value=True, key="sel_valueadd")
    sel_financing = st.checkbox("Financing & debt terms (incl. IO period)", value=True, key="sel_financing")
    sel_flags     = st.checkbox("Underwriting flags",                       value=True, key="sel_flags")
    sel_tax       = st.checkbox("Property tax & abatement",                 value=True, key="sel_tax")
    sel_repl      = st.checkbox("Replacement cost, insurance & management", value=True, key="sel_repl")

    st.markdown('<div class="rv-cb-divider"></div><div class="rv-panel-heading">Tab 2 — Comparables</div>', unsafe_allow_html=True)
    sel_rentcomps = st.checkbox("Garden & townhouse rent comps",            value=True, key="sel_rentcomps")
    sel_addinc2   = st.checkbox("Additional income opportunities",          value=True, key="sel_addinc2")
    sel_salecomps = st.checkbox("Sale comparables with buyer/seller",       value=True, key="sel_salecomps")

    st.markdown('<div class="rv-cb-divider"></div><div class="rv-panel-heading">Tab 3 — Demographics</div>', unsafe_allow_html=True)
    sel_addinc    = st.checkbox("Additional income",                        value=True, key="sel_addinc")
    sel_utilities = st.checkbox("Utilities & site information",             value=True, key="sel_utilities")
    sel_pop       = st.checkbox("Population & income (1-mi / 3-mi / 5-mi)",value=True, key="sel_pop")
    sel_afford    = st.checkbox("Affordability analysis",                   value=True, key="sel_afford")
    sel_schools   = st.checkbox("Schools, crime & quality of life",        value=True, key="sel_schools")
    sel_employers = st.checkbox("Major employers & economic drivers",       value=True, key="sel_employers")
    sel_market    = st.checkbox("Market, submarket & supply/demand",        value=True, key="sel_market")

    _n_sel = sum([sel_deal,sel_unitmix,sel_opstat,sel_valueadd,sel_financing,
                  sel_flags,sel_tax,sel_repl,sel_rentcomps,sel_addinc2,sel_salecomps,
                  sel_addinc,sel_utilities,sel_pop,sel_afford,sel_schools,sel_employers,sel_market])

    # Select all / Deselect all buttons
    _ba, _bb = st.columns(2)
    with _ba:
        if st.button("✓ Select All",   key="btn_selall",   use_container_width=True):
            for k in ["sel_deal","sel_unitmix","sel_opstat","sel_valueadd","sel_financing",
                      "sel_flags","sel_tax","sel_repl","sel_rentcomps","sel_addinc2","sel_salecomps",
                      "sel_addinc","sel_utilities","sel_pop","sel_afford","sel_schools","sel_employers","sel_market"]:
                st.session_state[k] = True
            st.rerun()
    with _bb:
        if st.button("✕ Deselect All", key="btn_deselall", use_container_width=True):
            for k in ["sel_deal","sel_unitmix","sel_opstat","sel_valueadd","sel_financing",
                      "sel_flags","sel_tax","sel_repl","sel_rentcomps","sel_addinc2","sel_salecomps",
                      "sel_addinc","sel_utilities","sel_pop","sel_afford","sel_schools","sel_employers","sel_market"]:
                st.session_state[k] = False
            st.rerun()

    st.markdown(f"""
<div class="rv-cb-divider"></div>
<div class="rv-panel-heading">Supported Brokers</div>
<div class="rv-brokers">JLL · CBRE · Marcus &amp; Millichap · Cushman &amp; Wakefield · Newmark · Colliers · Berkadia · Walker &amp; Dunlop · Northmarq</div>
<div class="rv-cb-divider"></div>
<div class="rv-meta">Processing time: <span>30–90 sec</span></div>
<div class="rv-meta">Max file size: <span>50 MB</span></div>
<div class="rv-meta">Output: <span>3-tab Excel (.xlsx)</span></div>
<div class="rv-meta" style="margin-top:8px;"><span style="color:#1DC9A4;font-weight:600;">{_n_sel} / 18</span> <span>sections selected</span></div>
""", unsafe_allow_html=True)

with main_col:
    st.markdown("""
<div style="padding: 36px 44px 0;">
  <div class="rv-section-label-wrap">
    <span class="rv-section-label">Upload Offering Memorandum</span>
    <div class="rv-section-label-line"></div>
  </div>
  <div class="rv-upload-card">
    <div class="rv-upload-title">Upload Your OM</div>
    <div class="rv-upload-sub">Supports any broker PDF — JLL, CBRE, Marcus &amp; Millichap, Northmarq and more</div>
  </div>
</div>
""", unsafe_allow_html=True)

    api_key = st.secrets.get("ANTHROPIC_API_KEY", os.environ.get("ANTHROPIC_API_KEY", ""))
    if not api_key:
        st.error("""
**API key not configured.**
- **Streamlit Cloud:** Go to ⚙️ Settings → Secrets → add: `ANTHROPIC_API_KEY = "sk-ant-..."`
- **Local:** Create `.streamlit/secrets.toml` with the same line.
""")
        st.stop()
    os.environ["ANTHROPIC_API_KEY"] = api_key

    # Pad the uploader to align with the upload card above
    _gap, _upload_col = st.columns([0.058, 0.942])
    with _upload_col:
        uploaded = st.file_uploader("Drop your OM PDF here", type=["pdf"], label_visibility="collapsed")

    if uploaded is None:
        st.markdown("""
<div style="padding: 0 44px;">
<div class="rv-steps">
  <div class="rv-step"><div class="rv-step-num">Step 1</div><div class="rv-step-txt">Upload a PDF using the box above</div></div>
  <div class="rv-step"><div class="rv-step-num">Step 2</div><div class="rv-step-txt">Click Analyze (takes 30–90 sec)</div></div>
  <div class="rv-step"><div class="rv-step-num">Step 3</div><div class="rv-step-txt">Download the underwriting Excel</div></div>
</div>
</div>
""", unsafe_allow_html=True)
        st.stop()

    size_mb = uploaded.size / 1024 / 1024
    st.markdown(f"""
<div style="padding: 0 44px;">
<div class="rv-file-info">
  📄 &nbsp;<b>{uploaded.name}</b> &nbsp;—&nbsp; {size_mb:.1f} MB
</div>
</div>
""", unsafe_allow_html=True)

    if st.button("🔍  Analyze Offering Memorandum", type="primary", use_container_width=True):
        progress_bar = st.progress(0, text="Starting...")
        status_box   = st.empty()

        def set_progress(pct, msg):
            progress_bar.progress(pct, text=msg)
            status_box.markdown(
                f"<div style='font-size:12px;color:#5A8FAA;margin-top:4px;'>{msg}</div>",
                unsafe_allow_html=True)

        try:
            set_progress(10, "Extracting text from PDF…")
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp.write(uploaded.getvalue())
                tmp_path = tmp.name
            pdf_text = extract_pdf_text(tmp_path)
            os.unlink(tmp_path)

            if not pdf_text or len(pdf_text.strip()) < 200:
                progress_bar.empty(); status_box.empty()
                st.error("Could not extract readable text. This PDF may be scanned — please use a text-based PDF.")
                st.stop()

            set_progress(30, f"Extracted {len(pdf_text):,} characters. Sending to Claude AI…")
            def log(msg): set_progress(55, msg)
            data = analyze_om(pdf_text, api_key, log)
            set_progress(75, "Generating Excel report…")
            sections = {
                "deal":      sel_deal,     "unitmix":   sel_unitmix,
                "opstat":    sel_opstat,   "valueadd":  sel_valueadd,
                "financing": sel_financing,"flags":     sel_flags,
                "tax":       sel_tax,      "repl":      sel_repl,
                "rentcomps": sel_rentcomps,"addinc2":   sel_addinc2,
                "salecomps": sel_salecomps,"addinc":    sel_addinc,
                "utilities": sel_utilities,"pop":       sel_pop,
                "afford":    sel_afford,   "schools":   sel_schools,
                "employers": sel_employers,"market":    sel_market,
            }
            excel_bytes = build_excel(data, uploaded.name, sections=sections)
            set_progress(100, "Done!")
            progress_bar.empty(); status_box.empty()

        except Exception as e:
            progress_bar.empty(); status_box.empty()
            st.error(f"**Error:** {e}")
            with st.expander("Full traceback"):
                import traceback; st.code(traceback.format_exc())
            st.stop()

        prop_name   = (data.get("property") or {}).get("name") or "Property"
        broker_name = (data.get("broker")   or {}).get("name") or "Unknown broker"

        st.markdown(f"""
<div style="padding: 0 44px;">
<div class="rv-success">
  <div class="rv-success-icon">✅</div>
  <div class="rv-success-text">Report ready &nbsp;·&nbsp; <b>{prop_name}</b> &nbsp;·&nbsp; Broker: {broker_name}</div>
</div>
</div>
""", unsafe_allow_html=True)

        safe = re.sub(r"[^a-zA-Z0-9_\- ]", "", prop_name).strip().replace(" ", "_")
        st.download_button(
            label="⬇️  Download Underwriting Report (.xlsx)",
            data=excel_bytes,
            file_name=f"{safe}_Underwriting_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        st.markdown("<div style='margin-top:24px;'>", unsafe_allow_html=True)
        pd_  = data.get("property")   or {}
        inv_ = data.get("investment") or {}
        va_  = data.get("value_add")  or {}
        tax_ = data.get("tax")        or {}
        m1, m2, m3, m4, m5, m6 = st.columns(6)
        with m1: st.metric("Units",       _v(pd_.get("units"), "n"))
        with m2: st.metric("Year Built",  _v(pd_.get("year_built")))
        with m3: st.metric("Occupancy",   _pct(pd_.get("occupancy_pct")))
        with m4: st.metric("Avg Rent",    _v(inv_.get("market_rent"), "$"))
        with m5: st.metric("Reno ROI",    _pct(va_.get("roi_pct")))
        with m6: st.metric("Tax Savings", _v(tax_.get("abatement_annual_savings"), "$"))
        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<div style='margin-top:20px;'>", unsafe_allow_html=True)
        tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
            "Unit Mix", "Value-Add", "Rent Comps", "Financials",
            "Tax & Abatement", "Demographics", "Financing", "Flags"
        ])

        with tab1:
            umix_ = data.get("unit_mix") or []
            if umix_:
                import pandas as pd
                st.dataframe(pd.DataFrame([{
                    "Type": u.get("type"), "Units": u.get("count"), "SF": u.get("sf"),
                    "Market Rent": u.get("market_rent"), "Eff. Rent": u.get("eff_rent"),
                    "Target Rent": u.get("target_rent"), "Upside/Unit": u.get("upside"),
                } for u in umix_]), use_container_width=True, hide_index=True)
            else:
                st.info("No unit mix data extracted.")

        with tab2:
            import pandas as pd
            va_plans = va_.get("by_floor_plan") or []
            if va_plans:
                st.markdown('<div class="gold-header">Floor Plan Value-Add</div>', unsafe_allow_html=True)
                st.dataframe(pd.DataFrame([{
                    "Type": p.get("type"), "SF": p.get("sf"), "Units": p.get("units"),
                    "In-Place Rent": p.get("inplace_rent"), "Rehab Cost": p.get("rehab_cost"),
                    "Premium/Unit": p.get("premium"), "Post-Rehab Rent": p.get("post_rehab_rent"),
                } for p in va_plans]), use_container_width=True, hide_index=True)
            levers_ = data.get("value_add_levers") or []
            if levers_:
                st.markdown('<div class="gold-header">Revenue Upside Levers</div>', unsafe_allow_html=True)
                st.dataframe(pd.DataFrame([{
                    "Lever": lv.get("lever"), "Units": lv.get("units"),
                    "Mo. Premium": lv.get("monthly_premium"), "Annual Upside": lv.get("annual_upside"),
                    "Notes": lv.get("notes"),
                } for lv in levers_]), use_container_width=True, hide_index=True)
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Reno Cost", _v(va_.get("total_cost"), "$"))
                st.metric("Cost Per Unit",   _v(va_.get("cost_per_unit"), "$"))
            with col2:
                st.metric("Annual Premium",  _v(va_.get("annual_premium"), "$"))
                st.metric("Monthly Premium", _v(va_.get("monthly_premium"), "$"))
            with col3:
                st.metric("ROI",             _pct(va_.get("roi_pct")))
                st.metric("Exterior CapEx",  _v(va_.get("exterior_capex"), "$"))

        with tab3:
            import pandas as pd
            rg_     = data.get("rent_comps_garden")    or []
            rth_    = data.get("rent_comps_townhouse") or []
            rcomps_ = data.get("rent_comps")           or []
            if rg_:
                st.markdown('<div class="gold-header">Garden Comps</div>', unsafe_allow_html=True)
                st.dataframe(pd.DataFrame([{"Property": c.get("name"), "Rent": c.get("rent"), "Notes": c.get("notes")} for c in rg_]),
                             use_container_width=True, hide_index=True)
            if rth_:
                st.markdown('<div class="gold-header">Townhouse Comps</div>', unsafe_allow_html=True)
                st.dataframe(pd.DataFrame([{"Property": c.get("name"), "Rent": c.get("rent"), "Notes": c.get("notes")} for c in rth_]),
                             use_container_width=True, hide_index=True)
            if rcomps_:
                st.markdown('<div class="gold-header">Full Comp Detail</div>', unsafe_allow_html=True)
                st.dataframe(pd.DataFrame([{
                    "Property": c.get("name"), "Type": c.get("comp_type"),
                    "Built": c.get("year_built"), "Units": c.get("units"),
                    "Occ": _pct(c.get("occupancy")), "Mkt Rent": _v(c.get("total_market"), "$"),
                    "Avg SF": c.get("avg_sf"),
                } for c in rcomps_]), use_container_width=True, hide_index=True)
            if not rg_ and not rth_ and not rcomps_:
                st.info("No rent comp data extracted.")

        with tab4:
            fin_ = data.get("financials") or {}
            periods_ = fin_.get("periods") or []
            inc_     = fin_.get("income_lines") or []
            exp_     = fin_.get("expense_lines") or []
            noi_     = fin_.get("noi") or {}
            if periods_ and (inc_ or exp_):
                import pandas as pd
                rows = []
                for line in inc_ + exp_:
                    row = {"Line Item": line.get("item")}
                    for p in periods_:
                        row[p] = _v(line.get("values", {}).get(p), "$")
                    rows.append(row)
                if noi_:
                    row = {"Line Item": "NET OPERATING INCOME"}
                    for p in periods_:
                        row[p] = _v(noi_.get(p), "$")
                    rows.append(row)
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
            else:
                st.info("No financial data extracted.")

        with tab5:
            tax_ = data.get("tax") or {}
            import pandas as pd
            rows = [(k.replace("_"," ").title(), v) for k, v in tax_.items() if v and k != "abatement_program"]
            if rows:
                st.markdown('<div class="gold-header">Property Tax Detail</div>', unsafe_allow_html=True)
                st.dataframe(pd.DataFrame(rows, columns=["Field","Value"]), use_container_width=True, hide_index=True)
            if tax_.get("abatement_program"):
                st.markdown('<div class="gold-header">Tax Abatement Program</div>', unsafe_allow_html=True)
                st.info(tax_["abatement_program"])
            if not rows and not tax_.get("abatement_program"):
                st.info("No tax data extracted.")

        with tab6:
            demo_ = data.get("demographics") or {}
            import pandas as pd
            cols_d = ["population_3mi","population_5mi","hh_income_3mi","hh_income_5mi",
                      "median_age","renter_pct","college_pct","white_collar"]
            demo_rows = [(k.replace("_"," ").title(), demo_.get(k)) for k in cols_d if demo_.get(k)]
            if demo_rows:
                st.dataframe(pd.DataFrame(demo_rows, columns=["Metric","Value"]), use_container_width=True, hide_index=True)
            else:
                st.info("No demographics data extracted.")

        with tab7:
            fin_i_ = data.get("financing") or {}
            nf_i_  = fin_i_.get("new_financing")  or {}
            asd_i_ = fin_i_.get("assumable_debt") or {}
            import pandas as pd
            rows_f = []
            if asd_i_.get("lender"):
                rows_f += [
                    ("Type",          "Assumable Debt"),
                    ("Lender",        asd_i_.get("lender")),
                    ("Loan Amount",   _v(asd_i_.get("loan_amount"), "$")),
                    ("Interest Rate", _pct(asd_i_.get("interest_rate"))),
                    ("Rate Type",     asd_i_.get("rate_type")),
                    ("IO Period",     asd_i_.get("interest_only_period")),
                    ("Maturity Date", asd_i_.get("maturity_date")),
                ]
            if nf_i_.get("lender") or nf_i_.get("loan_type"):
                rows_f += [
                    ("Type",          "New Financing"),
                    ("Lender",        nf_i_.get("lender")),
                    ("Loan Amount",   _v(nf_i_.get("loan_amount"), "$")),
                    ("Interest Rate", _pct(nf_i_.get("interest_rate"))),
                    ("LTV",           _pct(nf_i_.get("loan_to_value"))),
                    ("IO Period",     nf_i_.get("interest_only_period")),
                ]
            if rows_f:
                st.dataframe(pd.DataFrame([(k,v) for k,v in rows_f if v],
                             columns=["Field","Value"]), use_container_width=True, hide_index=True)
            else:
                st.info(f"Offering type: {fin_i_.get('offering_type') or 'Not specified'}")

        with tab8:
            flags_ = data.get("flags") or []
            if flags_:
                for fl in flags_:
                    cat  = (fl.get("category") or "").lower()
                    cls  = "flag-warn"   if cat in ("warning","caution","risk") else \
                           "flag-good"   if cat in ("opportunity","upside") else \
                           "flag-verify" if cat == "verify" else "flag-info"
                    icon = "⚠️" if cls=="flag-warn" else "\u2705" if cls=="flag-good" else "\U0001f50d" if cls=="flag-verify" else "\u2139️"
                    st.markdown(f"""
<div class="{cls}">
  <div class="flag-title">{icon} &nbsp;{fl.get('title','')}</div>
  <div class="flag-body">{fl.get('detail','')}</div>
</div>""", unsafe_allow_html=True)
            else:
                st.info("No underwriting flags extracted.")

        st.markdown("</div>", unsafe_allow_html=True)
